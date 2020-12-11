"""
################################################################################################################
###                                                 排课算法:                                                 ###
################################################################################################################
    算法框架：
    1. 排课顺序： 班级-科目下 手排课，对开连堂，对开自由，连堂课 - 关联班级连堂课，自由课 - 关联班级自由课的顺序排课
    2. 交换解决冲突：当自由课存在冲突时：
                    1次交换：与班级内已排位置交换
                    2次交换：班级内已排位置交换后与班级内已排位置2次交换
    3. 最后，如果自由课还存在冲突，记录冲突位置 -> conflict_course_list
       采用，solve_conflict.py 进行深度交换，成功则停止计算 / 失败则进入下一个循环，达到 max iteration 停止
    输出结果：
    是否解决所有冲突 True -- 表示最终课表没有冲突点; False -- 表示存在冲突点
    min conflict number:3 深度交换前的迭代结果中最小冲突数
    排课结果: [[['数', '政', '语', '英', None, '化', '化', '不排'] ..
################################################################################################################
"""
import random
import openpyxl
import copy
import gc
import time
import functools
import sys
import json
# from . import data_source as data_source
# import data_source as data_source
import algorithmapp.algorithm.assign_schedule.data_source as data_source


# debug
# import algorithmapp.algorithm.assign_schedule.data_g2_demo as data
# import algorithmapp.algorithm.assign_schedule.test_assign_schedule as assign_schedule
# from test_assign_schedule import main


class ClsWeekSection(object):
    def __init__(self, cls, week, section):
        self.cls = cls
        self.week = week
        self.section = section

    def __eq__(self, other):
        if isinstance(other, self.__class__):
            return (self.cls == other.cls) and (self.week == other.week) and (self.section == other.section)

    def __hash__(self):
        return hash(self.cls + self.week + self.section)

    pass


class AssignSchedule:
    """
     ################################################################################################################
     01: initialize

        初始化用户输入参数：
        course_hours： 记录各个科目需要排课的总课时数
        free_course_hours: 记录安排完固定课程后，非连堂课需要的总课时数
        linked_course_hours_classes: 各个班级所需排连堂课课时数
        courses: 所有的科目
        flow_classes_time: 走班课时间
        manual_assign_time: 手排课程时间
        linked_course_times: 连堂课可排时间
        linked_course_hours: 连堂课的标准所需课时数
        class_num: 班级对应的数字号, eg {'1班'：1, ...}
        class_amount: 班级数
        subj_no_assign: 固定科目不排课
        classes_name: 班级名称
        static_week: 周上课天数
        static_section: 日课时数
        teacher_no_assign: 固定老师不排课
        conflict_course_list: 记录所有不可交换点, 当自由课、上午最大课时存在冲突点时，记录冲突位置 --> 返回排课结果，
                                               当对开、连堂、连锁、课程存在冲突点 --> 重排
        重要参数：
        classes_subj: 记录classes_subj【班级】【科目】是否已排
        day_classes: 记录某课程在某天是否已排 day_classes[cla][day][subj] == 1
        schedules: 记录所有班级的课表 schedules[cla][day][section] = 'subj'
        linked_course_time_all_classes: 班级可排连堂课容器 —— 记录所有班级的连堂课可排时间 初始化
            = linked_course_times - manual_assign_time - flow_classes_time
        free_position_all_classes： 班级可排位置容器 —— 记录固定课程、连堂课排完后，所有班级剩余课程（以下统称：free 课程）的可排时间
        already_assigned_free_positions: 已排free课程的位置， 用于与冲突交换
        already_assigned_linked_positions：已排连堂课位置， 用于冲突点交换
     ################################################################################################################
    """

    def __init__(self, data):
        self.data = data
        # class_hours to record how many classes left
        self.course_hours = self.data.COURSE_TOTAL
        # 　5/26 -01 新增数据结构：早上最大课时数 教师互斥 首末最大课时数  某节课最多介个班同时上课==============bug:字典赋值，修改1个2个都改
        self.course_morning_hours = copy.deepcopy(self.data.COURSE_MORNING_TOTAL)
        self.conflict_teachers = self.data.CONFLICT_TEACHERS
        self.teacher_head_tail_num = copy.deepcopy(self.data.TEACHER_HEAD_TAIL_NUM)
        self.course_same_time_num = copy.deepcopy(self.data.COURSE_SAME_TIME_NUM)
        # ========================================================
        self.free_course_hours = []
        self.linked_course_hours_classes = []
        self.class_course_teacher_info = self.data.CLASSES_COURSE_TEACHER
        self.courses = self.data.COURSE
        self.classes = self.data.CLASSES
        self.morning_class_num = self.data.MORNING_CLASS_NUM
        self.flow_classes_time = self.data.FLOW_CLASS_TIME
        self.not_assign_time = self.data.NOT_ASSIGN_TIME
        self.manual_assign_time = self.data.PART_SCHEDULE
        self.linked_course_times = self.data.LINK_COURSE_TIMES
        self.linked_course_hours = self.data.COURSE_LINKED_HOURS
        # class_num:班级名称list class_amount:班级数量
        # CLASSES_NUM ={“1班”：1，“2班”：2， ... ， “9班”：9}
        self.class_num = self.data.CLASSES_NUM
        # CLASS_NUM = 9 一共有9个班
        self.class_amount = self.data.CLASS_NUM
        self.subj_no_assign = self.data.NOT_ASSIGN_COURSE
        self.classes_name = self.data.CLASSES
        self.static_week = self.data.WEEK
        self.static_section = self.data.SECTION
        self.teacher_no_assign = self.data.NOT_ASSIGN_TEACHER
        self.conflict_course_list = []
        self.locked_cla_course = self.data.LOCKED_CLA_COURSE
        self.link_course_count = self.data.LINK_COURSE_COUNT_PEER_DAY
        #####################################增加变量 by Mayuan########################################
        self.linked_course_flag = self.data.LINKED_COURSE_FLAG
        self.class_rest_pos = [2, 0]
        ###############################################################################################
        # 6/30 排课失败追踪指标
        self.morning_count_failed = 0
        # define classes_subj : whether class has finished assignment subj
        # [{"语":0, "数":0, "英":0, "物":0, "化":0, "生":0},{}，{}，{} ...{...}]
        self.classes_subj = []
        # CLASS_NUM = 9 一共有9个班
        for i in range(self.data.CLASS_NUM):
            cla = {}
            for subj in self.class_course_teacher_info[i + 1]:
                cla[subj] = 0
            self.classes_subj.append(cla)
        self.day_classes = []
        # self.classes_linked_subj = []
        # class_amount = 9
        for i in range(self.class_amount):
            # self.linked_course_hours_classes.append(copy.deepcopy(self.linked_course_hours))
            # 5/26 -02 新增不同班级不同标准课时数
            self.linked_course_hours_classes.append(copy.deepcopy(self.linked_course_hours[i + 1]))
        for i in range(self.data.CLASS_NUM):
            week = []
            # 之前range里硬编码的5
            for d in range(self.static_week):
                day = {}
                for subj in self.class_course_teacher_info[i + 1]:
                    day[subj] = 0
                week.append(day)
            self.day_classes.append(week)
        self.schedules = []
        for i in range(self.data.CLASS_NUM):
            week = []
            # 之前range里硬编码的5
            for d in range(self.static_week):
                day = []
                for j in range(self.data.SECTION):
                    day.append(None)
                week.append(day)
            self.schedules.append(week)
        # define linked_course_times for each class:
        # 初始化的时候移除了已排课程
        self.linked_course_time_all_classes = []
        # class_amount = 9 9个班
        # 145-169是根据手排课的安排，把各个班本来可以连堂的时间点剔除，即，当连堂和手排冲突，剔除连堂
        for m in range(self.class_amount):
            # 初始化每个班可排连堂的时间点
            linked_course_time_a_classes = copy.deepcopy(self.linked_course_times)
            # 151-163把某个班所有的 潜在可排连堂时间 根据 手排课要求 进行潜在可排连堂时间的剔除
            # linked_course_time_a_classes - fixed schedule
            for index in range(len(linked_course_time_a_classes) - 1, -1, -1):
                link = linked_course_time_a_classes[index]
                flag = False
                # 把该班对应的所有
                for assigned in self.manual_assign_time:
                    # bug 11: index with num
                    # class_num是班级和班级代号的映射集合
                    # ==号右侧代表当前班级，==号左侧代表手排课中的班级
                    if self.class_num[assigned['class']] - 1 == m:
                        # bug 10: 连堂课排两节，都不能与已排课程冲突
                        if (assigned['week'] == link['week'] and assigned['section'] == link['section']) or \
                                (assigned['week'] == link['week'] and assigned['section'] == link['section'] + 1):
                            linked_course_time_a_classes.pop(index)
                            flag = True
                            break
                #  假如这个班没有手排课的要求，就去判断是否该潜在可排连堂时间是留给走课的
                if flag:
                    continue
                for flow in self.flow_classes_time:
                    if flow['week'] == link['week'] and flow['section'] == link['section'] or \
                            flow['week'] == link['week'] and flow['section'] == link['section'] + 1:
                        linked_course_time_a_classes.pop(index)
                        break
            self.linked_course_time_all_classes.append(linked_course_time_a_classes)
        # define free positions for each class:
        self.free_position_all_classes = []
        # define free already assigned positions:
        self.already_assigned_free_positions = []
        # define linked already assigned positions:
        self.already_assigned_linked_positions = []
        for cla in range(self.class_amount):
            self.already_assigned_free_positions.append([])
            self.already_assigned_linked_positions.append([])

    """
    02 - 01: Assign locked courses
    
    02 - 02: Assign fixed classes:
        a. 安排固定课程到课表中，包括：
           走班课程，手排课
        b. 手排课安排完成后，更新连堂课剩余的课时数，用来安排后续的连堂课
    """

    def assign_fixed(self):
        # assign flow classes:
        # FLOW_CLASS_TIME = [{'week': 2, 'section': 5},]
        for time in self.flow_classes_time:
            # classes = ['1班', '2班', '3班', '4班', '5班', '6班', '7班', '8班', '9班']
            for cla in range(len(self.classes)):
                # zero-based
                day = time['week'] - 1
                section = time['section'] - 1
                # FLOW_SIGN = "走"
                self.schedules[cla][day][section] = self.data.FLOW_SIGN
        # assign manual assigned classes:
        # manual_assign_time = [{'week': 1, 'section': 3, 'class': '1班', 'course': '12'}, {'week': 2, 'section': 3, 'class': '1班', 'course': '12'}]
        # manual_time = {'week': 1, 'section': 3, 'class': '1班', 'course': '12'}
        for manual_time in self.manual_assign_time:
            # class_name = "1班"
            class_name = manual_time['class']
            # bug3: cla_num get real cla -1 to index
            # class_num = {"1班":1，"2班":2，"3班":3}
            cla = self.class_num[class_name] - 1
            day = manual_time['week'] - 1
            section = manual_time['section'] - 1
            course = manual_time['course']
            # class_course_teacher_info 中的班级表示班级号的数字
            teacher = self.class_course_teacher_info[cla + 1][course]
            # 5/26 8.上午最大课时数：手排课
            # morning_class_num = 这个学校上午要上几节课，全校统一
            if section <= self.morning_class_num - 1:
                # course_morning_hours: cla 表示班级 1 代表1班
                # course_morning_hours=COURSE_MORNING_TOTAL = {'1班': {'语': 6, '数': 6, '英': 5, '物': 4, '化': 4, '生': 4}}
                if cla + 1 in self.course_morning_hours and course in self.course_morning_hours[cla + 1]:
                    self.course_morning_hours[cla + 1][course] -= 1
            # 5/26, 6/30 11.教师首末节最大课时数: 手排课
            if not self.whether_section_satisfy_start_end_num(section, teacher):
                return False
            # 5/26 9.同节班数: 手排课
            if course in self.course_same_time_num and not self.whether_course_same_time_num(course, day, section):
                print(f"手排课{course}数量超出同节班数限制，请重新安排手排课！")
                return False
            self.schedules[cla][day][section] = course
            # update status of day and subject:
            self.day_classes[cla][day][course] = 1
            # 这里，如果该课程只排了一部分，比如体育课只排1节，不能把该科目去掉不排
            # self.classes_subj[cla][course] = 1
        # update classes_linked_course_hours:
        for i in range(len(self.manual_assign_time)):
            # class_num = {"1班":1，"2班":2，"3班":3}
            # manual_assign_time = [{'week': 1, 'section': 3, 'class': '1班', 'course': '12'},]
            cla = self.class_num[self.manual_assign_time[i]['class']] - 1
            course = self.manual_assign_time[i]['course']
            if 'link' in self.manual_assign_time[i]:
                if self.manual_assign_time[i]['link']:
                    self.linked_course_hours_classes[cla][course] -= 1
        # 额外安排的课程linked_hours = 0
        # for cla in self.extra_assign_time:
        #     cla_num = self.class_num[cla]
        #     for unit in self.extra_assign_time[cla]:
        #         course = unit['course']
        #         self.linked_course_hours_classes[cla_num - 1][course] = 0
        # assign not assign classes:
        for time in self.not_assign_time:
            for cla in range(len(self.classes)):
                day = time['week'] - 1
                section = time['section'] - 1
                self.schedules[cla][day][section] = self.data.NOT_ASSIGN_SIGN

    def assign_locked_course_linked(self):
        self.update_linked_positions()
        # 每个班的索引
        for cla in range(self.class_amount):
            # class_course_teacher_info={1: {'语': '毕姗', '数': '曹宇', '英': '路平', '物': '王碧雪', '化': '刘臣'},}
            # subj 这个班所有的课
            for subj in self.class_course_teacher_info[cla + 1]:
                # classes = ['1班', '2班', '3班', '4班', '5班', '6班', '7班', '8班', '9班', '10班']
                # key 是由两个for循环拼成的，双层for循环主要是为了拼成所有key的组合
                key = str(self.classes[cla]) + '-' + subj
                related_cla = self.get_related_class(cla, subj)
                # locked_cla_course = {'1班-化': {'class': '2班', 'subject': '生'}, '10班-政': {'class': '7班', 'subject': '地'}}
                if key in self.locked_cla_course:
                    cla_lock = self.locked_cla_course[key]['class']
                    # 获取对开班级和科目：class_num = {“1班”：1，“2班”：2， ... ， “9班”：9}
                    # cla_lock = 拿到cla_lock的数字索引
                    cla_lock = self.class_num[cla_lock] - 1
                    # 对开课程
                    subj_lock = self.locked_cla_course[key]['subject']
                    related_cla_lock = self.get_related_class(cla_lock, subj_lock)
                    # Assign subj for cla meanwhile assign subj_lock for cla_lock:
                    linked_hours = self.linked_course_hours_classes[cla][subj]
                    linked_hours_locked = self.linked_course_hours_classes[cla_lock][subj_lock]
                    hours = min(linked_hours, linked_hours_locked)
                    cla_time_linked = self.linked_course_time_all_classes[cla]
                    # 本科目连堂潜在课位集合
                    subj_time_linked = copy.deepcopy(cla_time_linked)
                    cla_time_locked_linked = self.linked_course_time_all_classes[cla_lock]
                    # 对开科目潜在连堂课集合
                    subj_time_locked_linked = copy.deepcopy(cla_time_locked_linked)
                    # 约束条件增加位置1：
                    # linked_course： subj_time - subj_no_assign
                    self.remove_subj_no_assign(subj, subj_time_linked)
                    # 移除所有固定课程等关联班级已排冲突位置： 教师冲突
                    self.remove_already_related_conflict_course_position_linked(subj, cla,
                                                                                subj_time_linked)
                    # linked_course： subj_time - teacher_no_assign
                    self.remove_teacher_no_assign(subj, cla, subj_time_linked)
                    # 移除所有已经排了2天连堂课的天数： # 当当天连堂课已经超过了最大限制，则pop掉这天的这个潜在连堂课位
                    self.remove_already_assigned_two_linked_classes(cla, subj_time_linked)
                    # Also remove subj_time_linked_locked:
                    # -----------------------------------------------------------------------------------
                    # linked_course subj_time - subj_no_assign
                    self.remove_subj_no_assign(subj_lock, subj_time_locked_linked)
                    # 移除所有固定课程等关联班级已排冲突位置： 教师冲突
                    self.remove_already_related_conflict_course_position_linked(subj_lock, cla_lock,
                                                                                subj_time_locked_linked)
                    # linked_course subj_time - teacher_no_assign
                    self.remove_teacher_no_assign(subj_lock, cla_lock, subj_time_locked_linked)
                    # 移除所有已经排了2天连堂课的天数：
                    self.remove_already_assigned_two_linked_classes(cla_lock, subj_time_locked_linked)
                    # 5/26 13.教师互斥：对开连堂课
                    self.remove_conflict_teacher_position_linked(cla, subj, subj_time_locked_linked)
                    # Two locked subj 取交集：
                    both_ok_position = []
                    for grid in subj_time_linked:
                        if grid in subj_time_locked_linked:
                            both_ok_position.append(grid)
                    # BOTH_OK_POSITION 移除最后一个位置
                    for j in range(len(both_ok_position) - 1, -1, -1):
                        if both_ok_position[j]['section'] == self.static_section:
                            both_ok_position.pop(j)
                    for i in range(hours):
                        if len(both_ok_position) == 0:
                            return False

                        # 5/26 8.上午最大课时数：对开连堂课
                        # day，section是随机选出来的符合条件的对开连堂课位
                        day, section = self.get_random_linked_position_locked(cla, subj, both_ok_position, cla_lock,
                                                                              subj_lock, 2)
                        if day is None:
                            return False
                        # Assign class to schedule:把对开连堂课安排到schedule里
                        self.schedules[cla][day][section] = subj
                        self.schedules[cla][day][section + 1] = subj
                        self.schedules[cla_lock][day][section] = subj_lock
                        self.schedules[cla_lock][day][section + 1] = subj_lock
                        # day_classes是用来判断这门课今天排没排的
                        self.day_classes[cla][day][subj] = 1
                        self.day_classes[cla_lock][day][subj] = 1
                        # 排课后，约束条件增加位置2： remove adjacent day position
                        # self.remove_linked_next_day(subj, cla, both_ok_position, day + 1)
                        # after assign update all available linked course time in cla:
                        # day and section - index
                        for k in range(len(cla_time_linked) - 1, -1, -1):
                            value = cla_time_linked[k]
                            if (value['week'] - 1 == day and value['section'] - 1 == section) or \
                                    (value['week'] - 1 == day and value['section'] - 1 == section + 1) or \
                                    (value['week'] - 1 == day and value['section'] - 1 == section - 1):
                                cla_time_linked.pop(k)
                        for k in range(len(cla_time_locked_linked) - 1, -1, -1):
                            value = cla_time_locked_linked[k]
                            if (value['week'] - 1 == day and value['section'] - 1 == section) or \
                                    (value['week'] - 1 == day and value['section'] - 1 == section + 1) or \
                                    (value['week'] - 1 == day and value['section'] - 1 == section - 1):
                                cla_time_locked_linked.pop(k)
                        for j in range(len(both_ok_position) - 1, -1, -1):
                            value = both_ok_position[j]
                            if value['week'] - 1 == day:
                                both_ok_position.pop(j)
                    # Update class:subj status:
                    self.classes_subj[cla][subj] = 1
                    self.classes_subj[cla_lock][subj_lock] = 1
        return True

    def assign_locked_course_free(self):
        self.get_all_free_position()
        self.update_all_course_hours()
        # init status: class_subj = 0
        # bug 6: not all equals to 0, manual schedule should stay 1:
        for i in range(len(self.classes_subj)):
            for subj in self.classes_subj[i]:
                self.classes_subj[i][subj] = 0
        for cla in range(self.class_amount):
            free_position_cla = self.free_position_all_classes[cla]
            # class_course_teacher_info = {
            #     '1': {'语': '毕姗', '数': '曹宇', '英': '路平', '物': '王碧雪', '化': '刘臣'},
            # }
            for subj in self.class_course_teacher_info[cla + 1]:
                # key="1班-语"
                key = str(self.classes[cla]) + '-' + subj
                related_cla = self.get_related_class(cla, subj)
                if key in self.locked_cla_course:
                    # cla_lock and cla should be index
                    cla_lock = self.class_num[self.locked_cla_course[key]['class']] - 1
                    subj_lock = self.locked_cla_course[key]['subject']
                    free_position_cla_lock = self.free_position_all_classes[cla_lock]
                    related_cla_lock = self.get_related_class(cla_lock, subj_lock)
                    if self.free_course_hours[cla][subj] == 0 or self.free_course_hours[cla_lock][subj_lock] == 0:
                        continue
                    if self.classes_subj[cla][subj] == 1 or self.classes_subj[cla_lock][subj_lock] == 1:
                        continue
                    related_conflict_position = []
                    # 这个班当前的剩余可排课位
                    free_position_subj = copy.deepcopy(self.free_position_all_classes[cla])
                    # 对开班级剩余可排课位
                    free_position_subj_lock = copy.deepcopy(self.free_position_all_classes[cla_lock])
                    # 把该课禁止的课位ban掉
                    self.remove_subj_no_assign_free(subj, free_position_subj)
                    self.remove_subj_no_assign_free(subj_lock, free_position_subj_lock)
                    self.remove_teacher_no_assign_free(subj, cla, free_position_subj)
                    self.remove_teacher_no_assign_free(subj_lock, cla_lock, free_position_subj_lock)
                    # 挖去关联班级已排冲突的位置
                    self.remove_already_related_conflict_course_position(subj, cla, free_position_subj)
                    # self.remove_already_related_conflict_course_position(subj_lock, related_cla_lock,
                    #                                                      free_position_subj_lock)
                    self.remove_already_related_conflict_course_position(subj_lock, cla_lock,
                                                                         free_position_subj_lock)
                    # 5/26 13.教师互斥：对开自由课
                    self.remove_conflict_teacher_position_free(cla, subj, free_position_subj)
                    # add condition : bug 9 position_subj 挖去当天已排
                    # ###########感觉这块可以优化，这块的逻辑是今天如果已经排了手排课了，那么今天就不能再排这科课################
                    for l in range(len(free_position_subj) - 1, -1, -1):
                        value = free_position_subj[l]
                        if self.day_classes[cla][value['week']][subj] == 1:
                            free_position_subj.pop(l)
                    for l in range(len(free_position_subj_lock) - 1, -1, -1):
                        value = free_position_subj_lock[l]
                        if self.day_classes[cla][value['week']][subj] == 1:
                            free_position_subj_lock.pop(l)
                    # ##############################################################################################
                    """
                        连锁班级可排位置 然后取交集
                    """
                    both_ok_position = []
                    for grid in free_position_subj:
                        if grid in free_position_subj_lock:
                            both_ok_position.append(grid)
                    hours = min(self.free_course_hours[cla][subj], self.free_course_hours[cla_lock][subj_lock])
                    for hour in range(hours):
                        # if conflict: swap with already assigned
                        if len(both_ok_position) == 0:
                            print("No locked course position!")
                            return False
                        # assign to free position:
                        # time = random.randint(0, len(both_ok_position) - 1)
                        # day = both_ok_position[time]['week']
                        # section = both_ok_position[time]['section']
                        # related_conflict_position.append([day, section])
                        # 5/26: 8.上午最大课时数：对开自由课
                        day, section = self.get_random_free_position_locked(cla, subj, both_ok_position, cla_lock,
                                                                            subj_lock, 1)
                        if day is None:
                            return False
                        related_conflict_position.append([day, section])
                        # assign subj to cla schedule:
                        self.schedules[cla][day][section] = subj
                        self.schedules[cla_lock][day][section] = subj_lock
                        # after assign update all available free course time in cla:
                        for k in range(len(free_position_cla) - 1, -1, -1):
                            value = free_position_cla[k]
                            if value['week'] == day and value['section'] == section:
                                free_position_cla.pop(k)
                                break
                        for k in range(len(free_position_cla_lock) - 1, -1, -1):
                            value = free_position_cla_lock[k]
                            if value['week'] == day and value['section'] == section:
                                free_position_cla_lock.pop(k)
                                break
                        for j in range(len(both_ok_position) - 1, -1, -1):
                            value = both_ok_position[j]
                            if value['week'] == day:
                                both_ok_position.pop(j)
                        # update status:
                        self.day_classes[cla][day][subj] = 1
                        self.day_classes[cla_lock][day][subj_lock] = 1
                    self.classes_subj[cla][subj] = 1
                    self.classes_subj[cla_lock][subj_lock] = 1
        return True

    """
    # 03: Assign linked courses:
        安排连堂课：
        temp_cla_linked_times:  记录班级可排连堂课的时间，每次排完课后update
        temp_subj_linked_times: 记录科目可排连堂课的时间，初始化：周可排（已-固定位置不排）- 固定科目不排 - 固定教师不排
                                排课后，temp_subj_linked_times - 当天已排 - 相邻天位置（remove_linked_next_day）
        安排顺序：
            遍历班级：
                遍历科目：
                    遍历科目课时：
                        ...
                安排关联班级同一科目：
    """

    def assign_linked_courses(self):
        self.update_linked_positions()
        for cla in range(self.class_amount):
            # update linked_course_time_all_classes[cla]
            temp_cla_linked_times = self.linked_course_time_all_classes[cla]
            math_linked_day = []
            for subj in self.class_course_teacher_info[cla + 1]:
                related_cla = self.get_related_class(cla, subj)
                # Already assigned:
                if self.classes_subj[cla][subj] == 1:
                    continue
                linked_course_num = self.linked_course_hours_classes[cla][subj]
                # No linked_hours left:
                if linked_course_num == 0:
                    continue
                # Assign class 1 subj from "语":
                temp_subj_linked_times = copy.deepcopy(temp_cla_linked_times)
                # 约束条件增加位置1：
                # linked_course subj_time - subj_no_assign
                self.remove_subj_no_assign(subj, temp_subj_linked_times)
                # 移除所有固定课程等关联班级已排冲突位置： 教师冲突
                self.remove_already_related_conflict_course_position_linked(subj, cla, temp_subj_linked_times)
                # linked_course subj_time - teacher_no_assign
                self.remove_teacher_no_assign(subj, cla, temp_subj_linked_times)
                # 移除所有已经排了2天连堂课的天数：
                self.remove_already_assigned_two_linked_classes(cla, temp_subj_linked_times)
                # 5/26 13. 教师互斥：连堂课
                self.remove_conflict_teacher_position_linked(cla, subj, temp_subj_linked_times)
                related_conflict_position = []
                for i in range(linked_course_num):
                    if len(temp_subj_linked_times) == 0:
                        print("Start swap linked course position!")
                        if self.swap_conflict_linked(temp_cla_linked_times, cla, subj, temp_subj_linked_times):
                            continue
                        else:
                            print("No linked course swap positions.path.dirname(os.path.dirname(__file__))ion!")
                            return False
                    # 随机选择可排课程位置：
                    # time = random.randint(0, len(temp_subj_linked_times) - 1)
                    # day = temp_subj_linked_times[time]['week'] - 1
                    # section = temp_subj_linked_times[time]['section'] - 1
                    # 5/26: 8.上午最大课时数：连堂课
                    day, section = self.get_random_linked_position(cla, subj, temp_subj_linked_times, 2)
                    if day is None:
                        return False
                    # add to related conflict pos:
                    related_conflict_position.append([day, section - 1])
                    related_conflict_position.append([day, section])
                    related_conflict_position.append([day, section + 1])
                    # assign subj to cla schedule: \ section and day are index
                    self.schedules[cla][day][section] = subj
                    self.schedules[cla][day][section + 1] = subj
                    # add to math_linked_day
                    if subj == '数':
                        math_linked_day.append(day + 1)
                    # format: real date
                    self.already_assigned_linked_positions[cla].append({'week': day + 1, 'section': section + 1})
                    # update class:day:subj status:
                    self.day_classes[cla][day][subj] = 1
                    # 排课后，约束条件增加位置2： remove adjacent day position
                    # self.remove_linked_next_day(subj, cla, temp_subj_linked_times, day + 1)
                    # after assign update all available linked course time in cla:
                    # day and section - index
                    for k in range(len(temp_cla_linked_times) - 1, -1, -1):
                        value = temp_cla_linked_times[k]
                        if (value['week'] - 1 == day and value['section'] - 1 == section) or \
                                (value['week'] - 1 == day and value['section'] - 1 == section + 1) or \
                                (value['week'] - 1 == day and value['section'] - 1 == section - 1):
                            temp_cla_linked_times.pop(k)
                    for j in range(len(temp_subj_linked_times) - 1, -1, -1):
                        value = temp_subj_linked_times[j]
                        if value['week'] - 1 == day:
                            temp_subj_linked_times.pop(j)
                # Update class:subj status:
                self.classes_subj[cla][subj] = 1
                # Assign subj in related class(current cla):
                if not self.assign_related_class(related_cla, subj, related_conflict_position, math_linked_day):
                    return False
        return True

    # 03 - 00: update linked course positions: 已经排好课的位置，以及排好课前面的位置都不能排连堂课
    # def update_linked_positions(self):
    #     # class_amount = 此次排课活动有几个班级
    #     for cla in range(self.class_amount):
    #         i = len(self.linked_course_time_all_classes[cla]) - 1
    #         # 遍历每个班的潜在连堂课位
    #         while i >= 0:
    #             # cla：day：section：zero based； i：每个班潜在连堂课的索引
    #             day = self.linked_course_time_all_classes[cla][i]['week'] - 1
    #             section = self.linked_course_time_all_classes[cla][i]['section'] - 1
    #             # 在初始化潜在连堂课位的时候就没有存储最后一个section，所以如果最后一个有课，sec 6 不会被删除
    #             if self.schedules[cla][day][section] is not None:
    #                 # 这里需要修改，需要根据‘大课间是否算连堂’进行处理
    #                 self.linked_course_time_all_classes[cla].pop(i)
    #                 if i - 1 >= 0 and section != 2:
    #                     self.linked_course_time_all_classes[cla].pop(i - 1)
    #                     i -= 1
    #             #         static_week = 一周有几天
    #             elif self.schedules[cla][day][section] is None and \
    #                     section + 1 < self.static_week and \
    #                     self.schedules[cla][day][section + 1] is not None:
    #                 self.linked_course_time_all_classes[cla].pop(i)
    #             # remove section == 7  static_section = 一天有几节课
    #             elif section == self.static_section - 1:
    #                 self.linked_course_time_all_classes[cla].pop(i)
    #             i -= 1

    # 2020/09/08 created by Mayuan
    def update_linked_positions(self):
        # 这里面装的全是已排课程的对象ClsWeekSection()
        scheduled_cls_list = []
        for cls in range(len(self.schedules)):
            for day in range(len(self.schedules[cls])):
                for section in range(len(self.schedules[cls][day])):
                    # all zero based
                    if self.schedules[cls][day][section] is not None:
                        scheduled_cls = ClsWeekSection(cls, day, section)
                        scheduled_cls_list.append(scheduled_cls)
        # 根据已排课程，往set里面加入潜在的要删除的时间点对象ClsWeekSection()
        deleting_set = set()
        for scheduled_clss in scheduled_cls_list:
            cls = scheduled_clss.cls
            week = scheduled_clss.week
            section = scheduled_clss.section
            if section + 1 == self.morning_class_num or section == 0:
                # 只添加自己这一个删除位置
                deleting_obj = ClsWeekSection(cls, week, section)
                # 将元素 x 添加到集合 s 中，如果元素已存在，则不进行任何操作
                deleting_set.add(deleting_obj)
            else:
                # 添加自己和之前
                deleting_obj = ClsWeekSection(cls, week, section)
                deleting_obj_pre = ClsWeekSection(cls, week, section - 1)
                deleting_set.add(deleting_obj)
                deleting_set.add(deleting_obj_pre)
        #  把linked_course_time_all_classes中和deleting_set重合的点全部删除
        for scheduled_clzz in deleting_set:
            for cls in range(len(self.linked_course_time_all_classes)):
                for i in range(len(self.linked_course_time_all_classes[cls]) - 1, -1, -1):
                    if scheduled_clzz.section + 1 == self.linked_course_time_all_classes[cls][i][
                        'section'] and scheduled_clzz.week + 1 == self.linked_course_time_all_classes[cls][i][
                        'week'] and scheduled_clzz.cls == cls:
                        self.linked_course_time_all_classes[cls].pop(i)

    # 03 - 01: find related class:
    # return index:
    def get_related_class(self, cla, subj):
        related_cla = []
        # bug: cla + 1
        teacher = self.class_course_teacher_info[cla + 1][subj]
        for i in range(len(self.class_course_teacher_info)):
            if subj not in self.class_course_teacher_info[i + 1]:
                continue
            # bug2: i, cla are indexes
            if i != cla and self.class_course_teacher_info[i + 1][subj] == teacher:
                related_cla.append(i)
        return related_cla

    # 03 - 02: assign related class，subj：
    def assign_related_class(self, related_cla, subj, related_conflict_position, math_linked_day):
        for cla in related_cla:
            # check if subj already assigned:
            if self.classes_subj[cla][subj] == 1:
                continue
            linked_course_num = self.linked_course_hours_classes[cla][subj]
            temp_cla_linked_times = self.linked_course_time_all_classes[cla]
            temp_subj_linked_times = copy.deepcopy(self.linked_course_time_all_classes[cla])
            # 约束条件：
            # subj for cla need to remove all related position:
            self.remove_related_position(temp_subj_linked_times, related_conflict_position)
            # remove no assign class positions:
            self.remove_subj_no_assign(subj, temp_subj_linked_times)
            self.remove_teacher_no_assign(subj, cla, temp_subj_linked_times)
            # 移除所有固定课程等关联班级已排冲突位置： 教师冲突
            related_cla_cur = self.get_related_class(cla, subj)
            self.remove_already_related_conflict_course_position_linked(subj, cla, temp_subj_linked_times)
            # 移除所有已经排了2天连堂课的天数：
            self.remove_already_assigned_two_linked_classes(cla, temp_subj_linked_times)
            # 5/26 13. 教师互斥：连堂课关联班级
            self.remove_conflict_teacher_position_linked(cla, subj, temp_subj_linked_times)
            for i in range(linked_course_num):
                # Math linked course should be assigned on the same day
                if subj == '数':
                    if not self.remove_other_day(subj, cla, temp_subj_linked_times, math_linked_day):
                        return False
                if len(temp_subj_linked_times) == 0:
                    print("Start swap linked course")
                    if self.swap_conflict_linked(temp_cla_linked_times, cla, subj, temp_subj_linked_times):
                        continue
                    else:
                        print("No linked course swap position!")
                        return False
                # time = random.randint(0, len(temp_subj_linked_times) - 1)
                # day = temp_subj_linked_times[time]['week'] - 1
                # section = temp_subj_linked_times[time]['section'] - 1
                # 5/26: 8.上午最大课时数：连堂课关联班级
                day, section = self.get_random_linked_position(cla, subj, temp_subj_linked_times, 2)
                if day is None:
                    return False
                related_conflict_position.append([day, section])
                related_conflict_position.append([day, section + 1])
                # assign subj to cla schedule:
                self.schedules[cla][day][section] = subj
                self.schedules[cla][day][section + 1] = subj
                # remove adjacent day position:
                # self.remove_linked_next_day(subj, cla, temp_subj_linked_times, day + 1)
                # update class:day:subj status:
                self.day_classes[cla][day][subj] = 1
                # after assign update all available linked course time in cla:
                for k in range(len(temp_cla_linked_times) - 1, -1, -1):
                    value = temp_cla_linked_times[k]
                    if (value['week'] - 1 == day and value['section'] - 1 == section) or \
                            (value['week'] - 1 == day and value['section'] - 1 == section + 1) or \
                            (value['week'] - 1 == day and value['section'] - 1 == section - 1):
                        temp_cla_linked_times.pop(k)
                for j in range(len(temp_subj_linked_times) - 1, -1, -1):
                    value = temp_subj_linked_times[j]
                    if value['week'] - 1 == day:
                        temp_subj_linked_times.pop(j)
            # Update class:subj status:
            self.classes_subj[cla][subj] = 1
        return True

    # 03 - 02 - 01: remove related conflict position:
    def remove_related_position(self, temp_subj_linked_times, related_conflict_position):
        for i in range(len(temp_subj_linked_times) - 1, -1, -1):
            for j in range(len(related_conflict_position)):
                if temp_subj_linked_times[i]['week'] - 1 == related_conflict_position[j][0] and \
                        temp_subj_linked_times[i]['section'] - 1 == related_conflict_position[j][1]:
                    temp_subj_linked_times.pop(i)
                    break

    # add condition function:
    # 03 - 03 : subj_time - subj_no_assign_time / here free time is not index
    # 这个方法是为了获得每科潜在连堂课位
    def remove_subj_no_assign(self, subj, subj_time):
        # 遍历这一科的所有可能联排时间（这科全年级通用），减去该科被禁止的时间
        for i in range(len(subj_time) - 1, -1, -1):
            time = subj_time[i]
            # NOT_ASSIGN_COURSE = {
            #     '语': [{'week': 4, 'section': 5}, {'week': 4, 'section': 6}, {'week': 4, 'section': 7}, {'week': 4, 'section': 8},
            #           {'week': 3, 'section': 2}],}
            if subj in self.subj_no_assign:
                for no_assign_time in self.subj_no_assign[subj]:
                    # bug 12: linked course need to consider double position:
                    # if time['week'] == no_assign_time['week'] and time['section'] == no_assign_time['section']:
                    if (time['week'] == no_assign_time['week'] and time['section'] == no_assign_time['section']) or \
                            (time['week'] == no_assign_time['week'] and time['section'] + 1 == no_assign_time[
                                'section']):
                        subj_time.pop(i)
                        # self.update_all_course_hours()
                        break
            else:
                break

    # 03 - 04 : subj_time - teacher_no_assign_time
    def remove_teacher_no_assign(self, subj, cla, subj_time):
        teacher = self.class_course_teacher_info[cla + 1][subj]
        for i in range(len(subj_time) - 1, -1, -1):
            time = subj_time[i]
            if teacher in self.teacher_no_assign:
                for no_assign_time in self.teacher_no_assign[teacher]:
                    if (time['week'] == no_assign_time['week'] and time['section'] == no_assign_time['section']) or \
                            (time['week'] == no_assign_time['week'] and time['section'] + 1 == no_assign_time[
                                'section']):
                        subj_time.pop(i)
                        # self.update_all_course_hours()
                        break

    # IMO this method can be deleted because no one tells linked courses should be separated, created by MY
    # 03 - 05 : remove next linked position : day is not index
    def remove_linked_next_day(self, subj, cla, subj_time, day):
        next = day + 1
        pre = day - 1
        for i in range(len(subj_time) - 1, -1, -1):
            time = subj_time[i]
            if next <= self.static_week:
                if time['week'] == next:
                    subj_time.pop(i)
                    # self.update_all_course_hours()
                    continue
            if pre >= 1:
                if time['week'] == pre:
                    subj_time.pop(i)
                    # self.update_all_course_hours()

    # 03 - 06 : SWAP LINKED CONFLICT POINT - All date format are real date
    # 没考虑同一天连堂课交换
    # 空余位置与已排连堂课交换，free position 排除一天已排2连堂情况
    def swap_conflict_linked(self, free_position_cla, cla, subj, free_position_subj):
        # 如果当天连堂课 >=2 remove 这个free_position_cla 只移除已经安排的课程
        # self.remove_already_assigned_two_linked_classes(cla, free_position_cla)
        for i in range(len(free_position_cla)):
            # current position b
            position = free_position_cla[i]
            day_b = position['week']
            section_b = position['section']
            teacher_b = self.class_course_teacher_info[cla + 1][subj]
            subj_b = subj
            # 如果连堂课数量 >= 2 跳过 选取下一个位置
            # if already assigned course >= 2
            count = 0
            for j in range(self.static_section):
                if self.schedules[cla][day_b - 1][j] is not None and (j + 1 < self.static_section) and \
                        self.schedules[cla][day_b - 1][j] == self.schedules[cla][day_b - 1][j + 1]:
                    if self.schedules[cla][day_b - 1][j] != '走':
                        count += 1
            # 提取公共参数： 最大连堂课个数
            if count >= self.link_course_count:
                continue
            for j in range(len(self.already_assigned_linked_positions[cla])):
                # if already assigned position day already has subj:
                already_position = self.already_assigned_linked_positions[cla]
                if self.day_classes[cla][already_position[j]['week'] - 1][subj] == 1:
                    continue
                # get position to swap a:
                day_a = self.already_assigned_linked_positions[cla][j]['week']
                section_a = self.already_assigned_linked_positions[cla][j]['section']
                subj_a = self.schedules[cla][day_a - 1][section_a - 1]
                teacher_a = self.class_course_teacher_info[cla + 1][subj_a]
                # add condition: if day_b already has assigned subj_a or day_a already has assigned subj_b:
                # b: current free position a: already assigned position
                same_day = False
                if day_a == day_b:
                    same_day = True
                if not same_day and self.day_classes[cla][day_b - 1][subj_a] == 1 or self.day_classes[cla][day_a - 1][
                    subj_b] == 1:
                    continue
                # judge whether could swap:
                # if put b to a: check related class position b teacher is not teacher_b
                flag = True
                # bug 10: re_cla get teacher should plus 1
                # Get related class:
                relate_cla_a = self.get_related_class(cla, subj_a)
                relate_cla_b = self.get_related_class(cla, subj_b)
                for re_cla in relate_cla_b:
                    if (self.schedules[re_cla][day_a - 1][section_a - 1] == subj_b) or \
                            (self.schedules[re_cla][day_a - 1][section_a] == subj_b):
                        flag = False
                        break
                for re_cla in relate_cla_a:
                    # if put a to b:
                    if (self.schedules[re_cla][day_b - 1][section_b - 1] == subj_a) or \
                            (self.schedules[re_cla][day_b - 1][section_b] == subj_a):
                        flag = False
                        break
                if not flag:
                    continue
                # if a not in subj_no_assign position: day_a,section_a is not in subj_no_assign(subj_b)
                # 6/30 swap_conflit_linked中添加检测subj_b in self.subj_no_assign校验
                if subj_b in self.subj_no_assign:
                    for position in self.subj_no_assign[subj_b]:
                        if (position['week'] == day_a and position['section'] == section_a) or \
                                (position['week'] == day_a and position['section'] == section_a + 1):
                            flag = False
                            break
                    if not flag:
                        continue
                # if b not in subj_no_assign position: day_b,section_b is not in subj_no_assign(subj_a)
                if subj_a in self.subj_no_assign:
                    for position in self.subj_no_assign[subj_a]:
                        if position['week'] == day_b and position['section'] == section_b or \
                                position['week'] == day_b and position['section'] == section_b + 1:
                            flag = False
                            break
                    if not flag:
                        continue
                # after swap teacher_a not in teacher_no_assign:
                if teacher_a in self.teacher_no_assign:
                    for time in self.teacher_no_assign[teacher_a]:
                        if time['week'] == day_b and time['section'] == section_b or \
                                time['week'] == day_b and time['section'] == section_b + 1:
                            flag = False
                            break
                if not flag:
                    continue
                # after swap teacher_b not in teacher_no_assign:
                if teacher_b in self.teacher_no_assign:
                    for time in self.teacher_no_assign[teacher_b]:
                        if time['week'] == day_a and time['section'] == section_a or \
                                time['week'] == day_a and time['section'] == section_a + 1:
                            flag = False
                            break
                # After swap, no next_day with same course:
                pre_day_a = max(0, day_a - 1)
                next_day_a = min(self.static_week, day_a + 1)
                pre_day_b = max(0, day_b - 1)
                next_day_b = min(self.static_week, day_b + 1)
                # bug 15: 应该周边不能有连堂课 不是 不能有该科目
                if self.day_classes[cla][pre_day_a - 1][subj_b] == 1 or \
                        self.day_classes[cla][next_day_a - 1][subj_b] == 1:
                    flag = False
                if self.day_classes[cla][pre_day_b - 1][subj_a] == 1 or \
                        self.day_classes[cla][next_day_b - 1][subj_a] == 1:
                    flag = False
                if flag:
                    # swap in schedules and update status: a and b in cla
                    self.schedules[cla][day_b - 1][section_b - 1] = subj_a
                    self.schedules[cla][day_b - 1][section_b] = subj_a
                    self.schedules[cla][day_a - 1][section_a - 1] = subj_b
                    self.schedules[cla][day_a - 1][section_a] = subj_b
                    if not same_day:
                        self.day_classes[cla][day_b - 1][subj_a] = 1
                        self.day_classes[cla][day_b - 1][subj_b] = 0
                        self.day_classes[cla][day_a - 1][subj_b] = 1
                        self.day_classes[cla][day_a - 1][subj_a] = 0
                    # bug5: update already assigned:
                    self.already_assigned_linked_positions[cla].append({'week': day_b, 'section': section_b})
                    # update free position_cla and free position_subj:
                    for k in range(len(free_position_cla) - 1, -1, -1):
                        value = free_position_cla[k]
                        if value['week'] == day_b and value['section'] == section_b:
                            free_position_cla.pop(k)
                    for j in range(len(free_position_subj) - 1, -1, -1):
                        value = free_position_subj[j]
                        if value['week'] == day_a:
                            free_position_subj.pop(j)
                        if value['week'] == day_b and value['section'] == section_b:
                            free_position_subj.pop(j)
                    # self.update_all_course_hours()
                    return True
        # self.update_all_course_hours()
        return False

    # 03 - 07 : subj_time - subj_already_linked_course_position ( including pre and cur)
    # WARNINGS: 如果存在一个老师带两个班，教不同的科目，按照目前根据科目+老师找关联班级的方法，无法捕捉
    # 解决思路: 7/10 1.取消关联班级概念，分别排除本班以外的班级 2.冲突位置排查: 其它班级中只要教师相同即为冲突
    def remove_already_related_conflict_course_position_linked(self, subj, cla, subj_time):
        # def remove_already_related_conflict_course_position_linked(self, subj, related_cla, subj_time):
        # bug 15 迭代器
        for i in range(len(subj_time) - 1, -1, -1):
            # for re_cla in related_cla:
            teacher = self.class_course_teacher_info[cla + 1][subj]
            for re_cla in range(self.class_amount):
                if re_cla != cla:
                    week = subj_time[i]['week'] - 1
                    section = subj_time[i]['section'] - 1
                    # remove cur and pre
                    subj_cla = self.schedules[re_cla][week][section]
                    # 如果re_cla班级该位置为空，那么置教师为None
                    teacher_cla = None
                    if subj_cla in self.class_course_teacher_info[re_cla + 1]:
                        teacher_cla = self.class_course_teacher_info[re_cla + 1][subj_cla]
                    # if self.schedules[re_cla][week][section] == subj:
                    if teacher_cla == teacher:
                        subj_time.pop(i)
                        break
                    if section + 1 < self.static_section and \
                            self.schedules[re_cla][week][section + 1] in self.class_course_teacher_info[re_cla + 1] and \
                            self.class_course_teacher_info[re_cla + 1][
                                self.schedules[re_cla][week][section + 1]] == teacher:
                        subj_time.pop(i)
                        break

    #########################################################################################################
    # 03 - 07 - 02: remove_conflict_teacher_position_linked
    #        移除教师互斥的连堂课位置  5/26
    # INPUT: cla -- 班级序号; subj -- 当前科目; subj_time -- 当前班级科目可排位置
    # OUTPUT: 更新subj_time（移除教师互斥位置）
    # WARNINGS:
    # VERSION: 5/26
    # HISTORY:
    # 06/12/2020 STAR:CREATED
    #########################################################################################################
    def remove_conflict_teacher_position_linked(self, cla, subj, subj_time):
        # subj --> teacher: 找到这门课的老师
        teacher = self.class_course_teacher_info[cla + 1][subj]
        conflict_teacher = None
        # conflict_week = None
        # conflict_section = None
        conflict_positions = []
        # 获得互斥教师:
        # conflict_teachers = {'毕姗-杨青': [{'week': 2, 'section': 1},{'week': 3, 'section': 1},{'week': 4, 'section': 1}, {'week': 5, 'section': 1},{'week': 6, 'section': 1}]}
        for key in self.conflict_teachers:
            # key = '毕姗-杨青'
            if teacher in key:
                # keys = ['毕姗'，'杨青']
                keys = key.split('-')
                for ele in keys:
                    if ele != teacher:
                        conflict_teacher = ele
                # 6/30 不止一个冲突点[{'week': 4, 'section': 5},{'week': 5, 'section': 8}]
                # conflict_week = self.conflict_teachers[key]['week']
                # conflict_section = self.conflict_teachers[key]['section']
                # 6/30 conflict_positions={'week': 4, 'section': 5},{'week': 5, 'section': 8}
                conflict_positions = self.conflict_teachers[key]
                break
        for i in range(len(subj_time) - 1, -1, -1):
            week = subj_time[i]['week'] - 1
            section = subj_time[i]['section'] - 1
            # 每一个潜在的连堂课位
            position = {'week': week, 'section': section}
            # if week == conflict_week and section == conflict_section:
            # 6/30
            # conflict_positions =  [{'week': 2, 'section': 1},{'week': 3, 'section': 1},{'week': 4, 'section': 1},]
            if position in conflict_positions:
                # 遍历所有班级的互斥位置，排除教师互斥的位置
                for cla in range(self.class_amount):
                    # 映射出课程表位置的教师，判断是否为互斥教师
                    subj_cla = self.schedules[cla][week][section]
                    # 如果课表中位置为空:
                    teacher_cla = None
                    if subj_cla:
                        teacher_cla = self.class_course_teacher_info[cla + 1][subj_cla]
                    subj_cla_1 = self.schedules[cla][week][section + 1]
                    teacher_cla_1 = None
                    if subj_cla_1:
                        teacher_cla_1 = self.class_course_teacher_info[cla + 1][subj_cla_1]
                    if teacher_cla == conflict_teacher:
                        subj_time.pop(i)
                        break
                    if section + 1 < self.static_section and \
                            teacher_cla_1 == conflict_teacher:
                        subj_time.pop(i)
                        break

    # 03 - 08 : remove all other positions : day is not index
    def remove_other_day(self, subj, cla, subj_time, math_linked_day):
        for i in range(len(subj_time) - 1, -1, -1):
            time = subj_time[i]
            if time['week'] not in math_linked_day:
                subj_time.pop(i)
        if len(subj_time) == 0:
            return False
        return True

    # 03 - 09 : 移除所有已经排了2天连堂课的天数：
    # 当当天连堂课已经超过了最大限制，则pop掉这天的这个潜在连堂课位
    def remove_already_assigned_two_linked_classes(self, cla, subj_time):
        for i in range(len(subj_time) - 1, -1, -1):
            time = subj_time[i]
            # if already assigned course >= 2
            count = 0
            for j in range(self.static_section):
                if self.schedules[cla][time['week'] - 1][j] is not None and (j + 1 < self.static_section) and \
                        self.schedules[cla][time['week'] - 1][j] == self.schedules[cla][time['week'] - 1][j + 1]:
                    if self.schedules[cla][time['week'] - 1][j] != '走':
                        count += 1
            # 提取公共参数： 最大连堂课个数
            if count >= self.link_course_count:
                subj_time.pop(i)

    # 04: Assign free courses:
    # from here: day, section presents index in schedule
    def assign_free_courses(self):
        self.get_all_free_position()
        self.update_all_course_hours()
        # init status: class_subj = 0
        # bug 6: not all equals to 0, manual schedule should stay 1:
        for i in range(len(self.classes_subj)):
            for subj in self.classes_subj[i]:
                self.classes_subj[i][subj] = 0
        for cla in range(self.class_amount):
            free_position_cla = self.free_position_all_classes[cla]
            for subj in self.class_course_teacher_info[cla + 1]:
                related_cla = self.get_related_class(cla, subj)
                if self.classes_subj[cla][subj] == 1:
                    continue
                related_conflict_position = []
                free_position_subj = copy.deepcopy(self.free_position_all_classes[cla])
                self.remove_subj_no_assign_free(subj, free_position_subj)
                self.remove_teacher_no_assign_free(subj, cla, free_position_subj)
                # 挖去关联班级已排冲突的位置
                self.remove_already_related_conflict_course_position(subj, cla, free_position_subj)
                # 5/26 13.教师互斥:自由课
                self.remove_conflict_teacher_position_free(cla, subj, free_position_subj)
                # add condition : bug 9 position_subj 挖去当天已排
                for l in range(len(free_position_subj) - 1, -1, -1):
                    value = free_position_subj[l]
                    if self.day_classes[cla][value['week']][subj] == 1:
                        free_position_subj.pop(l)
                if self.free_course_hours[cla][subj] == 0:
                    continue
                for hours in range(self.free_course_hours[cla][subj]):
                    # if conflict: swap with already assigned
                    if len(free_position_subj) == 0:
                        print("Start swap")
                        if self.swap_conflict(free_position_cla, cla, subj, free_position_subj):
                            continue
                        else:
                            # 尝试交换以排好的两个点，多次判断是否可以交换
                            if not self.random_swap_already_assigned(free_position_cla, cla, subj, free_position_subj):
                                print("No swap position!")
                                self.conflict_course_list.append({cla: subj})
                                continue
                            else:
                                continue
                            # self.conflict_course_list.append({cla: subj})
                            # continue
                    # assign to free position:
                    # time = random.randint(0, len(free_position_subj) - 1)
                    # day = free_position_subj[time]['week']
                    # section = free_position_subj[time]['section']
                    # 5/26 8.上午最大课时数：自由课
                    day, section = self.get_random_free_position(cla, subj, free_position_subj, 1)
                    # 日期:5/26 3.课时间隔：自由课（如果课时数2-3 则排第一个课时，需要移除相邻天位置）
                    if hours == 0 and day is not None:  # 排第一个课时
                        self.remove_free_courses_next_day(cla + 1, day, subj, free_position_subj)
                    if day is None:
                        return False
                        # self.conflict_course_list.append({{cla: subj}})
                        # continue
                    related_conflict_position.append([day, section])
                    # assign subj to cla schedule:
                    self.schedules[cla][day][section] = subj
                    self.already_assigned_free_positions[cla].append({'week': day, 'section': section})
                    # after assign update all available free course time in cla:
                    for k in range(len(free_position_cla) - 1, -1, -1):
                        value = free_position_cla[k]
                        if value['week'] == day and value['section'] == section:
                            free_position_cla.pop(k)
                            break
                    for j in range(len(free_position_subj) - 1, -1, -1):
                        value = free_position_subj[j]
                        if value['week'] == day:
                            free_position_subj.pop(j)
                    # update status:
                    self.day_classes[cla][day][subj] = 1
                self.classes_subj[cla][subj] = 1
                # Assign subj in related class(current cla):
                if not self.assign_related_class_free(related_cla, subj, related_conflict_position):
                    return False
        return True

    # 04 - 01 :  get all free position: free position 中存index而不是天数
    def get_all_free_position(self):
        self.free_position_all_classes = []
        for cla in range(self.class_amount):
            # 每个班的free positions
            free_position_a_class = []
            for day in range(len(self.schedules[cla])):
                for section in range(len(self.schedules[cla][day])):
                    if self.schedules[cla][day][section] is None:
                        time = {'week': day, 'section': section}
                        free_position_a_class.append(time)
            self.free_position_all_classes.append(free_position_a_class)

    # 04 - 02 : update all course hours:
    # 2020/09/28，出现了排的课比设置的课时数多的情况，我的打算是在这里发现课时数超出的情况就pop掉 Mayuan
    def update_all_course_hours(self):
        self.free_course_hours = []
        # 一共有几个班
        for cla in range(self.class_amount):
            dic = {}
            for subj in self.class_course_teacher_info[cla + 1]:
                # dic[subj] = self.course_hours[subj]
                # 日期：5/26 course_hours 改为每个班级不同
                dic[subj] = self.course_hours[cla + 1][subj]
                # 遍历课表，出现该科目 dic[subj] - 1
                for day in range(self.static_week):
                    for section in range(self.static_section):
                        if self.schedules[cla][day][section] == subj:
                            dic[subj] -= 1
                            ######################################Mayuan
                            if dic[subj] < 0:
                                self.schedules[cla][day][section] = None
                                dic[subj] = 0
                            ######################################Mayuan
                # 额外安排课程的free_course_hours = 2:
                # for unit in self.extra_assign_time[self.classes[cla]]:
                #     if unit['course'] == subj:
                #         dic[subj] = unit['num']
            self.free_course_hours.append(dic)
        # for manual assigned subj in specific cla set to 0 :
        # for course in self.manual_assign_time:
        #     self.free_course_hours[self.class_num[course['class']] - 1][course['course']] = 0

    # 04 - 03: assign related free classes:
    def assign_related_class_free(self, related_cla, subj, related_conflict_position):
        for cla in related_cla:
            # check if subj already assigned:
            if self.classes_subj[cla][subj] == 1:
                continue
            related_class = self.get_related_class(cla, subj)
            free_course_num = self.free_course_hours[cla][subj]
            temp_cla_free_times = self.free_position_all_classes[cla]
            temp_subj_free_times = copy.deepcopy(self.free_position_all_classes[cla])
            # subj for cla need to remove all related position:
            self.remove_related_position_free(temp_subj_free_times, related_conflict_position)
            self.remove_subj_no_assign_free(subj, temp_subj_free_times)
            self.remove_teacher_no_assign_free(subj, cla, temp_subj_free_times)
            # 挖去关联班级已排的该课程位置
            self.remove_already_related_conflict_course_position(subj, cla, temp_subj_free_times)
            # 5/26 13.教师互斥:自由课关联班级
            self.remove_conflict_teacher_position_free(cla, subj, temp_subj_free_times)
            # add condition : subj_time 挖去当天已排
            for l in range(len(temp_subj_free_times) - 1, -1, -1):
                value = temp_subj_free_times[l]
                if self.day_classes[cla][value['week']][subj] == 1:
                    temp_subj_free_times.pop(l)
            for i in range(free_course_num):
                # if conflict:
                if len(temp_subj_free_times) == 0:
                    print("No related position left!")
                    if self.swap_conflict(temp_cla_free_times, cla, subj, temp_subj_free_times):
                        continue
                    else:
                        # 尝试交换以排好的两个点，多次判断是否可以交换
                        if not self.random_swap_already_assigned(temp_cla_free_times, cla, subj, temp_subj_free_times):
                            print("No swap position!")
                            self.conflict_course_list.append({cla: subj})
                            continue
                        else:
                            continue
                        # self.conflict_course_list.append({cla: subj})
                        # continue
                # time = random.randint(0, len(temp_subj_free_times) - 1)
                # day = temp_subj_free_times[time]['week']
                # section = temp_subj_free_times[time]['section']
                # 5/26 8.上午最大课时数：自由课关联班级
                day, section = self.get_random_free_position(cla, subj, temp_subj_free_times, 1)
                if day is None:
                    return False
                    # self.conflict_course_list.append({{cla: subj}})
                    # continue
                # 日期:5/26 3.课时间隔：自由课关联班级（如果课时数2-3 则排第一个课时，需要移除相邻天位置）
                if i == 0:  # 排第一个课时
                    self.remove_free_courses_next_day(cla + 1, day, subj, temp_subj_free_times)
                related_conflict_position.append([day, section])
                # assign subj to cla schedule:
                self.schedules[cla][day][section] = subj
                # update class:day:subj status:
                self.day_classes[cla][day][subj] = 1
                # update already assigned status:
                self.already_assigned_free_positions[cla].append({'week': day, 'section': section})
                # after assign update all available free course time in cla:
                for k in range(len(temp_cla_free_times) - 1, -1, -1):
                    value = temp_cla_free_times[k]
                    if value['week'] == day and value['section'] == section:
                        temp_cla_free_times.pop(k)
                for j in range(len(temp_subj_free_times) - 1, -1, -1):
                    value = temp_subj_free_times[j]
                    if value['week'] == day:
                        temp_subj_free_times.pop(j)
            # Update class:subj status:
            self.classes_subj[cla][subj] = 1
            self.update_all_course_hours()
        return True

    #########################################################################################################
    # 04 - 04 : swap_conflict:
    #        班级内交换解决冲突：当科目没有可排位置时，此函数假设科目暂时放在班级可排位置（遍历），并与班级已排位置比较和交换
    # INPUT: free_position_cla -- 当前班级可排位置; cla -- 班级序号; free_position_subj -- 当前班级科目可排位置;
    # OUTPUT: 返回是否交换成功
    # WARNINGS: * 如果与当天课程交换，不能是同一门 * 如果当天已有该课程， 不能交换
    # VERSION: 5/26
    # HISTORY:
    # 08/2019 STAR:CREATED
    #########################################################################################################
    def swap_conflict(self, free_position_cla, cla, subj, free_position_subj):
        for i in range(len(free_position_cla)):
            # current position b
            position = free_position_cla[i]
            day_b = position['week']
            section_b = position['section']
            teacher_b = self.class_course_teacher_info[cla + 1][subj]
            subj_b = subj
            for j in range(len(self.already_assigned_free_positions[cla])):
                # if already assigned position day already has subj:
                already_position = self.already_assigned_free_positions[cla]
                if self.day_classes[cla][already_position[j]['week']][subj] == 1:
                    continue
                # get position to swap a:
                day_a = self.already_assigned_free_positions[cla][j]['week']
                section_a = self.already_assigned_free_positions[cla][j]['section']
                subj_a = self.schedules[cla][day_a][section_a]
                teacher_a = self.class_course_teacher_info[cla + 1][subj_a]
                # if free_position == already_assigend position, continue
                if day_a == day_b and section_a == section_b:
                    continue
                # bug 19: 如果空余位置当天已经存在该课程，该位置无效
                # bug 20 free_position_cla 空余位置排除已经有了该课程的
                if self.day_classes[cla][day_a][subj_b] == 1 or self.day_classes[cla][day_b][subj_b] == 1:
                    continue
                if subj_a == subj_b:
                    continue
                # add condition: if day_b already has assigned subj_a or day_a already has assigned subj_b:
                same_day = False
                # bug 14： not consider day_a == day_b
                # add condition: if day_b already has assigned subj_a or day_a already has assigned subj_b:
                if day_a != day_b and (
                        self.day_classes[cla][day_b][subj_a] == 1 or self.day_classes[cla][day_a][subj_b] == 1):
                    continue
                # 2020/09/29
                # 判断是否冲突了教师互斥
                # 这里是不是应该加一个判断教师冲突的问题
                # ########################尝试########################2020/09/29 Mayuan##################################################
                conflict_teacher_flagg = False
                for conflict_teacher_keys in self.conflict_teachers:
                    if teacher_a in conflict_teacher_keys:
                        positions = self.conflict_teachers[conflict_teacher_keys]
                        for time_slot in positions:
                            if time_slot['week'] - 1 == day_b and time_slot['section'] - 1 == section_b:
                                conflict_teacher_flagg = True
                                break
                    if teacher_b in conflict_teacher_keys:
                        positions = self.conflict_teachers[conflict_teacher_keys]
                        for time_slot in positions:
                            if time_slot['week'] - 1 == day_a and time_slot['section'] - 1 == section_a:
                                conflict_teacher_flagg = True
                                break
                if conflict_teacher_flagg:
                    continue
                ########################################################Mayuan########################################################
                if day_a == day_b:
                    same_day = True
                # judge whether could swap:
                # if put b to a: check related class position b teacher is not teacher_b
                flag = True
                # bug 10: re_cla get teacher should plus 1
                # bug 13: 同一个班级下，不科目的关联班级不一样
                # related class:
                relate_cla_b = self.get_related_class(cla, subj_b)
                relate_cla_a = self.get_related_class(cla, subj_a)
                for re_cla in relate_cla_b:
                    if self.schedules[re_cla][day_a][section_a] == subj_b:
                        flag = False
                        break
                for re_cla in relate_cla_a:
                    # if put a to b:
                    if self.schedules[re_cla][day_b][section_b] == subj_a:
                        flag = False
                        break
                if not flag:
                    continue
                # if a not in subj_no_assign position: day_a,section_a is not in subj_no_assign(subj_b)
                if subj_b in self.subj_no_assign:
                    for position in self.subj_no_assign[subj_b]:
                        if position['week'] - 1 == day_a and position['section'] - 1 == section_a:
                            flag = False
                            break
                if not flag:
                    continue
                # if b not in subj_no_assign position: day_b,section_b is not in subj_no_assign(subj_a)
                if subj_a in self.subj_no_assign:
                    for position in self.subj_no_assign[subj_a]:
                        if position['week'] - 1 == day_b and position['section'] - 1 == section_b:
                            flag = False
                            break
                if not flag:
                    continue
                # after swap teacher_a not in teacher_no_assign:
                if teacher_a in self.teacher_no_assign:
                    for time in self.teacher_no_assign[teacher_a]:
                        if time['week'] - 1 == day_b and time['section'] - 1 == section_b:
                            flag = False
                            break
                if not flag:
                    continue
                # after swap teacher_b not in teacher_no_assign:
                if teacher_b in self.teacher_no_assign:
                    for time in self.teacher_no_assign[teacher_b]:
                        if time['week'] - 1 == day_a and time['section'] - 1 == section_a:
                            flag = False
                            break
                if flag:
                    # swap in schedules and update status: a and b in cla
                    self.schedules[cla][day_b][section_b] = subj_a
                    self.schedules[cla][day_a][section_a] = subj_b
                    if not same_day:
                        self.day_classes[cla][day_b][subj_a] = 1
                        self.day_classes[cla][day_b][subj_b] = 0
                        self.day_classes[cla][day_a][subj_b] = 1
                        self.day_classes[cla][day_a][subj_a] = 0
                    # bug5: update already assigned:
                    self.already_assigned_free_positions[cla].append({'week': day_b, 'section': section_b})
                    # update free position_cla and free position_subj:
                    for k in range(len(free_position_cla) - 1, -1, -1):
                        value = free_position_cla[k]
                        if value['week'] == day_b and value['section'] == section_b:
                            free_position_cla.pop(k)
                    for j in range(len(free_position_subj) - 1, -1, -1):
                        value = free_position_subj[j]
                        if value['week'] == day_a:
                            free_position_subj.pop(j)
                        if value['week'] == day_b and value['section'] == section_b:
                            free_position_subj.pop(j)
                    self.update_all_course_hours()
                    return True
        self.update_all_course_hours()
        return False

    # 04 - 05: remove related free conflict position:
    # bug 7: 2 system day and week --> index
    def remove_related_position_free(self, temp_subj_free_times, related_conflict_position):
        for i in range(len(temp_subj_free_times) - 1, -1, -1):
            for j in range(len(related_conflict_position)):
                if temp_subj_free_times[i]['week'] == related_conflict_position[j][0] and \
                        temp_subj_free_times[i]['section'] == related_conflict_position[j][1]:
                    temp_subj_free_times.pop(i)
                    break

    # 04 - 06 : subj_time - subj_no_assign_time / here free time is not index
    def remove_subj_no_assign_free(self, subj, subj_time):
        for i in range(len(subj_time) - 1, -1, -1):
            time = subj_time[i]
            if subj in self.subj_no_assign:
                for no_assign_time in self.subj_no_assign[subj]:
                    if time['week'] == no_assign_time['week'] - 1 and time['section'] == no_assign_time['section'] - 1:
                        subj_time.pop(i)

    # 04 - 07 : subj_time - teacher_no_assign_time
    def remove_teacher_no_assign_free(self, subj, cla, subj_time):
        teacher = self.class_course_teacher_info[cla + 1][subj]
        for i in range(len(subj_time) - 1, -1, -1):
            time = subj_time[i]
            if teacher in self.teacher_no_assign:
                for no_assign_time in self.teacher_no_assign[teacher]:
                    if time['week'] == no_assign_time['week'] - 1 and time['section'] == no_assign_time['section'] - 1:
                        try:
                            subj_time.pop(i)
                        except Exception:
                            print(subj_time)
                            print(i)

    # 04 - 08 : subj_time - subj_already_linked_course_position
    # WARNINGS: 如果存在一个老师带两个班，教不同的科目，按照目前根据科目+老师找关联班级的方法，无法捕捉
    # 解决思路: 7/10 1.关联班级概念放大，不管是否同科目，只要有相同的教师即为关联班级 2.冲突位置排查:其它班级中只要教师相同即为冲突
    def remove_already_related_conflict_course_position(self, subj, cla, subj_time):
        teacher = self.class_course_teacher_info[cla + 1][subj]
        for i in range(len(subj_time) - 1, -1, -1):
            for re_cla in range(self.class_amount):
                if re_cla != cla:
                    # subj_cla 中存储的是下标，非天数，可直接在schedules中读取课程
                    subj_cla = self.schedules[re_cla][subj_time[i]['week']][subj_time[i]['section']]
                    teacher_cla = None
                    if subj_cla in self.class_course_teacher_info[re_cla + 1]:
                        teacher_cla = self.class_course_teacher_info[re_cla + 1][subj_cla]
                    # if self.schedules[re_cla][subj_time[i]['week']][subj_time[i]['section']] == subj:
                    if teacher_cla == teacher:
                        subj_time.pop(i)
                        break

    ##################################################################################################################
    # 04 - 08 - 01 : remove_conflict_teacher_position_free
    #        遍历已排课表，移除当前班级科目可排位置中与其他班级教师互斥的位置
    # INPUT: cla -- 当前班级; subj -- 当前科目; subj -- 冲突科目; subj_time -- 当前班级科目可排位置;
    # OUTPUT: 移除教师互斥位置后班级科目可排位置subj_time(更新)
    # WARNINGS:
    # VERSION: 5/26
    # HISTORY:
    # 06/12/2020 STAR:CREATED
    ##################################################################################################################
    def remove_conflict_teacher_position_free(self, cla, subj, subj_time):
        teacher = self.class_course_teacher_info[cla + 1][subj]
        conflict_teacher = None
        # conflict_week = None
        # conflict_section = None
        conflict_positions = []
        # 获得互斥教师:
        for key in self.conflict_teachers:
            if teacher in key:
                keys = key.split('-')
                for ele in keys:
                    if ele != teacher:
                        conflict_teacher = ele
                # conflict_teacher = keys[1]
                # conflict_week = self.conflict_teachers[key]['week']
                # conflict_section = self.conflict_teachers[key]['section']
                conflict_positions = self.conflict_teachers[key]
                break
        for i in range(len(subj_time) - 1, -1, -1):
            # 5/26 去掉-1
            week = subj_time[i]['week']
            section = subj_time[i]['section']
            position = {'week': week, 'section': section}
            # if week == conflict_week and section == conflict_section:
            # 6/30
            if position in conflict_positions:
                # 遍历所有班级的互斥位置，排除教师互斥的位置
                for cla in range(self.class_amount):
                    subj_cla = self.schedules[cla][week][section]
                    teacher_cla = None
                    if subj_cla:
                        teacher_cla = self.class_course_teacher_info[cla + 1][subj_cla]
                    if teacher_cla == conflict_teacher:
                        subj_time.pop(i)
                        break

    ##################################################################################################################
    # 04 - 09: random_swap_already_assigned:
    #        交换班级内两个已排课程，再次查看是否可以通过swap_conflict解决科目冲突
    # INPUT: free_position_cla -- 当前班级可排位置; cla -- 班级序号; subj -- 冲突科目; free_position_subj -- 当前班级科目可排位置;
    # OUTPUT: 返回是否交换成功
    # WARNINGS:
    # VERSION: 5/26
    # HISTORY:
    # 08/2019 STAR:CREATED
    ##################################################################################################################
    def random_swap_already_assigned(self, free_position_cla, cla, subj, free_position_subj):
        # check if swap 2 positions whether could get a result:
        for i in range(len(self.already_assigned_free_positions[cla]) - 1):
            for j in range(i + 1, len(self.already_assigned_free_positions[cla])):
                # swap_two_positions:
                position_i = self.already_assigned_free_positions[cla][i]
                day_a = position_i['week']
                section_a = position_i['section']
                subj_a = self.schedules[cla][day_a][section_a]
                position_j = self.already_assigned_free_positions[cla][j]
                day_b = position_j['week']
                section_b = position_j['section']
                subj_b = self.schedules[cla][day_b][section_b]
                same_day = False
                if day_b == day_a:
                    same_day = True
                if self.check_swap(cla, subj_a, subj_b, day_a, section_a, day_b, section_b):
                    # swap:
                    # swap in schedules and update status: a and b in cla
                    self.schedules[cla][day_b][section_b] = subj_a
                    self.schedules[cla][day_a][section_a] = subj_b
                    if not same_day:
                        self.day_classes[cla][day_b][subj_a] = 1
                        self.day_classes[cla][day_b][subj_b] = 0
                        self.day_classes[cla][day_a][subj_b] = 1
                        self.day_classes[cla][day_a][subj_a] = 0
                    # After swap check if could swap:
                    if self.swap_conflict(free_position_cla, cla, subj, free_position_subj):
                        self.update_all_course_hours()
                        return True
                    else:
                        # Swap back: =============================================================
                        self.schedules[cla][day_b][section_b] = subj_b
                        self.schedules[cla][day_a][section_a] = subj_a
                        if not same_day:
                            self.day_classes[cla][day_b][subj_a] = 0
                            self.day_classes[cla][day_b][subj_b] = 1
                            self.day_classes[cla][day_a][subj_b] = 0
                            self.day_classes[cla][day_a][subj_a] = 1
        self.update_all_course_hours()
        return False

    # 04 - 10: 查看两点是否可交换 - bug 16 挖去当天已排位置 bug 19: 当天已存在该课程，不交换
    # 这里的day和section都是0索引的
    def check_swap(self, cla, subj_a, subj_b, day_a, section_a, day_b, section_b):
        # if free_position == already_assigend position, continue
        if day_a == day_b and section_a == section_b:
            return False
        # add condition: if day_b already has assigned subj_a or day_a already has assigned subj_b:
        same_day = False
        # bug 14： not consider day_a == day_b
        # add condition: if day_b already has assigned subj_a or day_a already has assigned subj_b:
        if day_a != day_b and (
                self.day_classes[cla][day_b][subj_a] == 1 or self.day_classes[cla][day_a][subj_b] == 1):
            return False
        # bug 19: else: 导致两个正常的点，被认为是同一天。交换后更新科目安排状态存在问题。
        if day_a == day_b:
            same_day = True
        # judge whether could swap:
        # if put b to a: check related class position b teacher is not teacher_b
        flag = True
        # bug 10: re_cla get teacher should plus 1
        # bug 13: 同一个班级下，不科目的关联班级不一样
        # related class:
        relate_cla_b = self.get_related_class(cla, subj_b)
        relate_cla_a = self.get_related_class(cla, subj_a)

        teacher_a = self.class_course_teacher_info[cla + 1][subj_a]
        teacher_b = self.class_course_teacher_info[cla + 1][subj_b]
        for re_cla in relate_cla_b:
            if self.schedules[re_cla][day_a][section_a] == subj_b:
                flag = False
                return False
        for re_cla in relate_cla_a:
            # if put a to b:
            if self.schedules[re_cla][day_b][section_b] == subj_a:
                flag = False
                return False
        # if a not in subj_no_assign position: day_a,section_a is not in subj_no_assign(subj_b)
        if subj_b in self.subj_no_assign:
            for position in self.subj_no_assign[subj_b]:
                if position['week'] - 1 == day_a and position['section'] - 1 == section_a:
                    flag = False
                    return False
        # if b not in subj_no_assign position: day_b,section_b is not in subj_no_assign(subj_a)
        if subj_a in self.subj_no_assign:
            for position in self.subj_no_assign[subj_a]:
                if position['week'] - 1 == day_b and position['section'] - 1 == section_b:
                    flag = False
                    return False
        # after swap teacher_a not in teacher_no_assign:
        if teacher_a in self.teacher_no_assign:
            for time in self.teacher_no_assign[teacher_a]:
                if time['week'] - 1 == day_b and time['section'] - 1 == section_b:
                    flag = False
                    return False
        # after swap teacher_b not in teacher_no_assign:
        if teacher_b in self.teacher_no_assign:
            for time in self.teacher_no_assign[teacher_b]:
                if time['week'] - 1 == day_a and time['section'] - 1 == section_a:
                    flag = False
                    return False
        # 2020/09/29 Mayuan
        # 这里是不是应该加一个判断教师冲突的问题
        # ########################尝试########################2020/09/29 Mayuan##################################################
        for conflict_teacher_keys in self.conflict_teachers:
            if teacher_a in conflict_teacher_keys:
                positions = self.conflict_teachers[conflict_teacher_keys]
                for time_slot in positions:
                    if time_slot['week'] - 1 == day_b and time_slot['section'] - 1 == section_b:
                        return False
            if teacher_b in conflict_teacher_keys:
                positions = self.conflict_teachers[conflict_teacher_keys]
                for time_slot in positions:
                    if time_slot['week'] - 1 == day_a and time_slot['section'] - 1 == section_a:
                        return False
        return flag

    ##################################################################################################################
    # 04 - 11: remove_free_courses_next_day
    #        课时间隔：每周2课时、3课时的课，需有一次两课时之间最少间隔一天（不能连着2天或3天上完）
    #        此函数，判断所排课程是否只有2-3个课时，如果是，移除第一节课相邻两天的位置，后面排课不和第一节课相邻
    # VERSION: 5/26
    # INPUT: cla -- 当前班级; day -- 天数-1（下标）; subj -- 科目; subj_time -- 当前班级科目可排位置;
    # OUTPUT: 更新移除相邻天后的班级科目可排位置
    # WARNINGS: day  表示下标 周五：4
    # HISTORY:
    # 06/12/2020 STAR:CREATED
    ##################################################################################################################
    def remove_free_courses_next_day(self, cla, day, subj, subj_time):
        # 判断课时数是否2-3：
        hours = self.course_hours[cla][subj]
        if hours >= 2 and hours <= 3:
            next = day + 1
            pre = day - 1
            for i in range(len(subj_time) - 1, -1, -1):
                time = subj_time[i]
                if next <= self.static_week:
                    if time['week'] == next:
                        subj_time.pop(i)
                        continue
                if pre >= 1:
                    if time['week'] == pre:
                        subj_time.pop(i)

    # 5 5/26 公共函数
    # 5-01-01 随机选择自由课排课位置
    def get_random_free_position(self, cla, subj, pool, option):
        # 如果尝试多次失败：返回False 重排
        count = 0
        day = None
        section = None
        while 1:
            count += 1
            if count > 200:
                return None, None
            time = random.randint(0, len(pool) - 1)
            # 下标
            day = pool[time]['week']
            section = pool[time]['section']
            # 5/26 9.同节班数：自由课
            if subj in self.course_same_time_num and not self.whether_course_same_time_num(subj, day, section):
                continue
            # 5/26 11.教师首末节最大课时数: 随机自由课满足教师首末节最大课时数
            teacher = self.class_course_teacher_info[cla + 1][subj]
            if not self.whether_section_satisfy_start_end_num(section, teacher):
                continue
            # Version:5/26 判断是否达到上午最多课时数
            if section <= self.morning_class_num - 1 and self.course_morning_hours[cla + 1][subj] >= option:
                self.course_morning_hours[cla + 1][subj] -= option
                break
            elif section > self.morning_class_num - 1:
                break
        return day, section

    # 5-01-02 随机选择连堂课排课位置
    def get_random_linked_position(self, cla, subj, pool, option):
        # 如果尝试多次失败：返回False 重排
        count = 0
        while 1:
            count += 1
            if count > 200:
                return None, None
            time = random.randint(0, len(pool) - 1)
            day = pool[time]['week'] - 1
            section = pool[time]['section'] - 1
            # 5/26 9.同节班数：连堂课
            if subj in self.course_same_time_num and not self.whether_course_same_time_num_linked(subj, day, section):
                continue
            # 5/26 11.教师首末节最大课时数: 随机连堂课也满足教师首末节最大课时数
            teacher = self.class_course_teacher_info[cla + 1][subj]
            if not self.whether_section_satisfy_start_end_num_linked(section, teacher):
                continue
            # Version:5/26 判断是否达到上午最多课时数
            if section < self.morning_class_num - 1 and self.course_morning_hours[cla + 1][subj] >= option:
                self.course_morning_hours[cla + 1][subj] -= option
                return day, section
            elif section >= self.morning_class_num:
                return day, section
        return day, section

    # 5-02-01 随机选择对开连堂课排课位置:(day)天数(因为pool中存储的为天数）
    def get_random_linked_position_locked(self, cla, subj, pool, cla_lock, subj_lock, option):
        count = 0
        day = None
        section = None
        while 1:
            count += 1
            if count > 200:
                return None, None
            time = random.randint(0, len(pool) - 1)
            day = pool[time]['week'] - 1
            section = pool[time]['section'] - 1
            # 5/26 9.同节班数：对开连堂课,如果同节班数不满足，直接continue
            if subj in self.course_same_time_num and not self.whether_course_same_time_num_linked(subj, day, section):
                continue
            if subj_lock in self.course_same_time_num and not self.whether_course_same_time_num_linked(subj_lock, day,
                                                                                                       section):
                continue
            # 5/26 11.教师首末节最大课时数: 对开连堂课也满足教师首末节最大课时数
            teacher = self.class_course_teacher_info[cla + 1][subj]
            teacher_lock = self.class_course_teacher_info[cla_lock + 1][subj_lock]
            # 上午首节连堂
            if section == 0:
                if teacher in self.teacher_head_tail_num:
                    if self.teacher_head_tail_num[teacher][0] <= 0:
                        print(f"手排课不满足教师{teacher}上午最大课时数限制，请重新排课！")
                        continue
                    self.teacher_head_tail_num[teacher][0] -= 1
                if teacher_lock in self.teacher_head_tail_num:
                    if self.teacher_head_tail_num[teacher_lock][0] <= 0:
                        print(f"手排课不满足教师{teacher_lock}上午最大课时数限制，请重新排课！")
                        continue
                    self.teacher_head_tail_num[teacher_lock][0] -= 1
            # 上午末节连堂
            # 张鑫 section == 2； 马原self.morning_class_num-2
            elif section == self.morning_class_num - 2:
                if teacher in self.teacher_head_tail_num:
                    if self.teacher_head_tail_num[teacher][1] <= 0:
                        print(f"手排课不满足教师{teacher}上午最大课时数限制，请重新排课！")
                        continue
                    self.teacher_head_tail_num[teacher][1] -= 1
                if teacher_lock in self.teacher_head_tail_num:
                    if self.teacher_head_tail_num[teacher_lock][1] <= 0:
                        print(f"手排课不满足教师{teacher_lock}上午最大课时数限制，请重新排课！")
                        continue
                    self.teacher_head_tail_num[teacher_lock][1] -= 1
            # 下午首节连堂
            # 张鑫 section == 4； 马原self.morning_class_num-2
            elif section == self.morning_class_num:
                if teacher in self.teacher_head_tail_num:
                    if self.teacher_head_tail_num[teacher][2] <= 0:
                        print(f"手排课不满足教师{teacher}上午最大课时数限制，请重新排课！")
                        continue
                    self.teacher_head_tail_num[teacher][2] -= 1
                if teacher_lock in self.teacher_head_tail_num:
                    if self.teacher_head_tail_num[teacher_lock][2] <= 0:
                        print(f"手排课不满足教师{teacher_lock}上午最大课时数限制，请重新排课！")
                        continue
                    self.teacher_head_tail_num[teacher_lock][2] -= 1
            # 下午末节连堂
            # 张鑫 section == 6； 马原self.static_section-2
            elif section == self.static_section - 2:
                if teacher in self.teacher_head_tail_num:
                    if self.teacher_head_tail_num[teacher][3] <= 0:
                        print(f"手排课不满足教师{teacher}下午最大课时数限制，请重新排课！")
                        continue
                    self.teacher_head_tail_num[teacher][3] -= 1
                if teacher_lock in self.teacher_head_tail_num:
                    if self.teacher_head_tail_num[teacher_lock][3] <= 0:
                        print(f"手排课不满足教师{teacher_lock}下午最大课时数限制，请重新排课！")
                        continue
                    self.teacher_head_tail_num[teacher_lock][3] -= 1
            # Version:5/26 判断是否达到上午最多课时数 对开课也满足上午最大课时数
            # option的意思是，连堂课需要在上午最大课时的基础上 -2
            if section < self.morning_class_num - 1 and self.course_morning_hours[cla + 1][subj] >= option and \
                    self.course_morning_hours[cla_lock + 1][
                        subj_lock] >= option:
                self.course_morning_hours[cla + 1][subj] -= option
                self.course_morning_hours[cla_lock + 1][subj_lock] -= option
                return day, section
            elif section >= self.morning_class_num:
                return day, section
        return day, section

    # 5-02-02 随机选择对开自由课排课位置：下标
    def get_random_free_position_locked(self, cla, subj, pool, cla_lock, subj_lock, option):
        count = 0
        day = None
        section = None
        while 1:
            count += 1
            if count > 200:
                return None, None
            time = random.randint(0, len(pool) - 1)
            day = pool[time]['week']
            section = pool[time]['section']
            # 5/26 9.同节班数：对开自由课
            if subj in self.course_same_time_num and not self.whether_course_same_time_num(subj, day, section):
                continue
            if subj_lock in self.course_same_time_num and not self.whether_course_same_time_num(subj_lock, day,
                                                                                                section):
                continue
            # 5/26 11.教师首末节最大课时数: 对开自由课也满足教师首末节最大课时数
            teacher = self.class_course_teacher_info[cla + 1][subj]
            teacher_lock = self.class_course_teacher_info[cla_lock + 1][subj_lock]
            if section == 0:
                if teacher in self.teacher_head_tail_num:
                    if self.teacher_head_tail_num[teacher][0] <= 0:
                        print(f"对开自由课位置不满足教师{teacher}上午首节最大课时数限制，请重新排课！")
                        continue
                    self.teacher_head_tail_num[teacher][0] -= 1
                if teacher_lock in self.teacher_head_tail_num:
                    if self.teacher_head_tail_num[teacher_lock][0] <= 0:
                        print(f"对开自由课位置不满足教师{teacher_lock}上午首节最大课时数限制，请重新排课！")
                        continue
                    self.teacher_head_tail_num[teacher_lock][0] -= 1
            elif section == 3:
                if teacher in self.teacher_head_tail_num:
                    if self.teacher_head_tail_num[teacher][1] <= 0:
                        print(f"对开自由课位置不满足教师{teacher}上午末节最大课时数限制，请重新排课！")
                        continue
                    self.teacher_head_tail_num[teacher][1] -= 1
                if teacher_lock in self.teacher_head_tail_num:
                    if self.teacher_head_tail_num[teacher_lock][1] <= 0:
                        print(f"对开自由课位置不满足教师{teacher_lock}上午末节最大课时数限制，请重新排课！")
                        continue
                    self.teacher_head_tail_num[teacher_lock][1] -= 1
            elif section == 4:
                if teacher in self.teacher_head_tail_num:
                    if self.teacher_head_tail_num[teacher][2] <= 0:
                        print(f"对开自由课位置不满足教师{teacher}下午首节最大课时数限制，请重新排课！")
                        continue
                    self.teacher_head_tail_num[teacher][2] -= 1
                if teacher_lock in self.teacher_head_tail_num:
                    if self.teacher_head_tail_num[teacher_lock][2] <= 0:
                        print(f"对开自由课位置不满足教师{teacher_lock}下午首节最大课时数限制，请重新排课！")
                        continue
                    self.teacher_head_tail_num[teacher_lock][2] -= 1
            elif section == 7:
                if teacher in self.teacher_head_tail_num:
                    if self.teacher_head_tail_num[teacher][3] <= 0:
                        print(f"手排课不满足教师{teacher}下午末节最大课时数限制，请重新排课！")
                        continue
                    self.teacher_head_tail_num[teacher][3] -= 1
                if teacher_lock in self.teacher_head_tail_num:
                    if self.teacher_head_tail_num[teacher_lock][3] <= 0:
                        print(f"手排课不满足教师{teacher_lock}下午末节最大课时数限制，请重新排课！")
                        continue
                    self.teacher_head_tail_num[teacher_lock][3] -= 1
            # Version:5/26 判断是否达到上午最多课时数 对开课也满足上午最大课时数
            if section <= self.morning_class_num - 1 and self.course_morning_hours[cla + 1][subj] >= option and \
                    self.course_morning_hours[cla_lock + 1][
                        subj_lock] >= option:
                self.course_morning_hours[cla + 1][subj] -= option
                self.course_morning_hours[cla_lock + 1][subj_lock] -= option
                break
            elif section > self.morning_class_num - 1:
                break
        return day, section

    # 5-03-01 计算同节班数是否满足约束条件，并更新同节班数
    def whether_course_same_time_num(self, subj, day, section):
        count = 0
        # 6/30 Bug: 需要-1 因为当前subj还没有排，选出位置后还要排subj
        max_num = self.course_same_time_num[subj] - 1
        #  # class_num = {"1班":1，"2班":2，"3班":3}
        for cla in range(self.class_amount):
            if self.schedules[cla][day][section] == subj:
                count += 1
        return count <= max_num

    # 5-03-02 计算同节班数是否满足约束条件，并更新同节班数
    def whether_course_same_time_num_linked(self, subj, day, section):
        count1 = 0
        count2 = 0
        max_num = self.course_same_time_num[subj] - 1
        for cla in range(self.class_amount):
            if self.schedules[cla][day][section] == subj:
                count1 += 1
            if self.schedules[cla][day][section + 1] == subj:
                count2 += 1
        return count1 <= max_num and count2 <= max_num

    # 5-04-01 自由课是否满足教师首末节课最大数量
    def whether_section_satisfy_start_end_num(self, section, teacher):
        if section == 0:
            if teacher in self.teacher_head_tail_num:
                if self.teacher_head_tail_num[teacher][0] <= 0:
                    print(f"排课位置不满足教师{teacher}上午首节最大课时数限制，请重新排课！")
                    return False
                self.teacher_head_tail_num[teacher][0] -= 1
        elif section == self.morning_class_num - 1:
            if teacher in self.teacher_head_tail_num:
                if self.teacher_head_tail_num[teacher][1] <= 0:
                    print(f"排课位置不满足教师{teacher}上午末节最大课时数限制，请重新排课！")
                    return False
                self.teacher_head_tail_num[teacher][1] -= 1
        elif section == self.morning_class_num:
            if teacher in self.teacher_head_tail_num:
                if self.teacher_head_tail_num[teacher][2] <= 0:
                    print(f"排课位置不满足教师{teacher}下午首节最大课时数限制，请重新排课！")
                    return False
                self.teacher_head_tail_num[teacher][2] -= 1
        elif section == self.static_section - 1:
            if teacher in self.teacher_head_tail_num:
                if self.teacher_head_tail_num[teacher][3] <= 0:
                    print(f"排课位置不满足教师{teacher}下午最大课时数限制，请重新排课！")
                    return False
                self.teacher_head_tail_num[teacher][3] -= 1
        return True

    # 5-04-02 连堂课是否满足教师首末节课最大数量
    def whether_section_satisfy_start_end_num_linked(self, section, teacher):
        if section == 0:
            if teacher in self.teacher_head_tail_num:
                if self.teacher_head_tail_num[teacher][0] <= 0:
                    print(f"排课位置不满足教师{teacher}上午首节最大课时数限制，请重新排课！")
                    return False
                self.teacher_head_tail_num[teacher][0] -= 1
        elif section == 2:
            if teacher in self.teacher_head_tail_num:
                if self.teacher_head_tail_num[teacher][1] <= 0:
                    print(f"排课位置不满足教师{teacher}上午末节最大课时数限制，请重新排课！")
                    return False
                self.teacher_head_tail_num[teacher][1] -= 1
        elif section == 4:
            if teacher in self.teacher_head_tail_num:
                if self.teacher_head_tail_num[teacher][2] <= 0:
                    print(f"排课位置不满足教师{teacher}下午首节最大课时数限制，请重新排课！")
                    return False
                self.teacher_head_tail_num[teacher][2] -= 1
        elif section == 6:
            if teacher in self.teacher_head_tail_num:
                if self.teacher_head_tail_num[teacher][3] <= 0:
                    print(f"排课位置不满足教师{teacher}下午最大课时数限制，请重新排课！")
                    return False
                self.teacher_head_tail_num[teacher][3] -= 1
        return True

    def output_schedules(self):
        # 将课表输出到excel中
        workbook = openpyxl.Workbook()
        for classes in self.classes_name:
            worksheet = workbook.create_sheet(classes)
            schedule = self.schedules[self.class_num[classes] - 1]
            # print(classes)
            for x in range(self.static_week):
                for y in range(self.static_section):
                    course = schedule[x][y]
                    if course is None:
                        continue
                    else:
                        value = classes + '-' + course
                        ## 5/26  因为每个班级的课程可能不同 所以取消self.courses,并用self.class_num[classes]替代
                        # if course in self.courses:
                        if course in self.class_course_teacher_info[self.class_num[classes]]:
                            value += ('-' + self.data.CLASSES_COURSE_TEACHER[self.class_num[classes]][course])
                        worksheet.cell(y + 1, x + 1, value)
        workbook.save('schedule.xlsx')

    def log_execution_time(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            start = time.perf_counter()
            iteration = 1
            for i in range(iteration):
                result = func(*args, **kwargs)
            end = time.perf_counter()
            print('{} took average {} ms'.format(func.__name__, (end - start) * 1000 / iteration))
            return result

        return wrapper

    # 主函数：
    @log_execution_time
    def run(self, data, solve_conflict):
        iteration = 0
        conflict = 0
        min_conflict = sys.maxsize
        ret_conflict_list = []
        print(min_conflict)
        result = None
        # 6/30 算法合理性检测 -- 0 班级数不全
        if self.course_hours.keys() != self.linked_course_hours.keys():
            res = f'连堂课或总课时中的班级数量不一致，请修改！'
            print(res)
            return res, None
        # 6/30 算法合理性检测 -- 1. 总课时数不超过一周最大课时数
        for key in self.course_hours.keys():
            course_hours_sum = 0
            for key_son in self.course_hours[key].keys():
                course_hours_sum += self.course_hours[key][key_son]
            #     之前张鑫大于号右边hardcode 40
            if course_hours_sum > self.data.WEEK * self.data.SECTION:
                res = f"{key}课时总数为{course_hours_sum},超出每周最大课时数限制！请重新设置课时数."
                print(f"{key}课时总数为{course_hours_sum},超出每周最大课时数限制！请重新设置课时数.")
                return res, None
        ## 6/30 算法合理性检测 -- 2. 科目手排课不超过科目总课时数 + 手排课没有重复位置
        part_schedule = {}  # init
        repeat_check = []
        # Init part_schedule 第一层: cla层 --（字典只能一层一层你赋值）
        for element in self.manual_assign_time:
            # 只能1-9班
            # cla = int(element['class'][0])
            cla = int(element['class'][0:-1])  # 马原

            part_schedule[cla] = {}
        # Init 手排课课程总数为
        for element in self.manual_assign_time:
            # 判断是否有重复:
            cla = int(element['class'][0:-1])  # 马原

            week = element['week']
            section = element['section']
            course = element['course']
            # 如果已存在,输出手排课重复
            if (cla, week, section) in repeat_check:
                res = f"手排课错误！{cla}班手排课:'{course}',位置:(week:{week},section:{section})安排重复，请修改！"
                print(res)
                return res, None
            repeat_check.append((cla, week, section))
            # 6/30 Bug: 1个班安排了多节课，这样会覆盖:
            part_schedule[cla][element['course']] = 0
        # Count 手排课中的课程总数
        for element in self.manual_assign_time:
            # 班级统一为数字：'1班'-1
            cla = int(element['class'][0:-1])  # 马原

            part_schedule[cla][element['course']] += 1
        # 判断手排课 + 连堂课是否超过总课时数:
        for cla in part_schedule.keys():
            res = None
            for course in part_schedule[cla].keys():
                # 检测cla course是否在course_hours, 否:报错
                if cla not in self.course_hours or course not in self.course_hours[cla]:
                    res = "手排课班级或科目不存在，请在班级总科目中录入！"
                    print(res)
                    return res
                if part_schedule[cla][course] + self.linked_course_hours[cla][course] * 2 > self.course_hours[cla][
                    course]:
                    res = f"{cla}班手排课'{course}'+连堂课课时总数超过课程要求的最大课时数，请修改手排课数量！"
                    print(res)
                    return res, None
        # 6/30 算法合理性检测 -- 3. 连排课课时数数不超过总课时数
        for cla in self.linked_course_hours.keys():
            for course in self.linked_course_hours[cla]:
                if course in self.course_hours[cla]:
                    if self.linked_course_hours[cla][course] * 2 > self.course_hours[cla][course]:
                        res = f"{cla}班连堂课'{course}'课时数超过课程要求的最大课时数限制，请修改！"
                        print(res)
                        return res, None
        # 6/30 算法合理性检测 -- 4. 科目一周排课节数 <= 5 (包括连堂课，因为同科目一天不能排两次）
        # 另外: 如果通过增加连堂课节数能修正用户操作，则增加连堂课节数
        for cla in self.linked_course_hours.keys():
            count_course_week = 0
            for course in self.linked_course_hours[cla]:
                if course in self.course_hours[cla]:
                    count_course_week = self.course_hours[cla][course] - self.linked_course_hours[cla][course]
                    if count_course_week > 5:
                        modify_flag = False
                        i = self.linked_course_hours[cla][course]
                        while i < 4:
                            i += 1
                            if self.course_hours[cla][course] - i <= 5:
                                self.linked_course_hours[cla][course] = i
                                modify_flag = True
                                break
                        if modify_flag:
                            continue
                        res = f"{cla}班课程{course}总节数（连堂课算1节）超过5节，不满足每天只排一节课（包括连堂课）需求！请修改课程数量！"
                        print(res)
                        return res, None
        # 6/30 算法合理性检测 -- 5. 每个班连堂课总数<LINK_COURSE_COUNT_PEER_DAY*5 否则无法安排完所有的连堂课
        for cla in self.course_hours.keys():
            count_cla_link = 0
            for subj in self.course_hours[cla]:
                count_cla_link += self.linked_course_hours[cla][subj]
            if count_cla_link > self.link_course_count * 5:
                res = f"因为已设置每天最大安排{self.link_course_count}节连堂课，{cla}班设置的连堂课总节数超出上限！请修改连堂课数量或者上限！"
                print(res)
                return res, None
        count_failed = 0
        # 6/30 设置失败原因追踪，排课失败后给用户反馈失败原因 -- 暂时反馈上午最大课时数失败
        while iteration <= 15:
            gc.collect()
            self.__init__(data)
            print('iteration:{}'.format(iteration))
            # 01: Assign fixed flow classes： -1
            self.assign_fixed()
            # 如果条件过于严格，无法排出课表，提示用户修改排课条件:
            if count_failed > 30:
                res = '排课条件过于严格，无法找到符合条件的结果，请修改排课条件!比如增加课程上午可排最大课时数或者减少科目连堂课数量.'
                print(res)
                return res, None
            # 02: Assign locked courses: {'1班-化': {'class': '2班', 'subject': '生'}, '9班-政': {'class': '7班', 'subject': '地'}}
            # 对开联排课                                             对开自由课
            if (not self.assign_locked_course_linked()) or (not self.assign_locked_course_free()):
                print("assign locked course failed")
                count_failed += 1
                continue
            # 03: Assign linked classes
            if not self.assign_linked_courses():
                print("assign linked course failed")
                count_failed += 1
                continue
            # 04: Assign reminded(free) classes:
            if not self.assign_free_courses():
                print("assign free courses failed")
                count_failed += 1
                continue
            print(self.schedules)
            conflict = len(self.conflict_course_list)
            for i in range(len(self.conflict_course_list)):
                print(self.conflict_course_list[i], len(self.conflict_course_list))
            # 如果上面的都过了，说明已经没有严重的错误了，是可以输出的课表

            # 解决冲突
            # auther:ice
            conflict_list = []
            for a_conflict in self.conflict_course_list:
                for key, value in a_conflict.items():
                    hava_flag = False
                    for class_name in self.class_num:
                        if self.class_num[class_name] == key:
                            hava_flag = True
                            break
                    if hava_flag:
                        conflict_list.append({'class': class_name, 'course': value})
            solve_conflict_flag = False
            # 对排课结果进行深度搜索： 并更新self.schedules
            if solve_conflict is not None:
                solve_conflict_flag = solve_conflict(self.schedules, conflict_list)
                conflict = len(conflict_list)
            print('是否解决所有冲突', solve_conflict_flag)
            # if solve_conflict_flag:
            #     return self.schedules, None
            if conflict < min_conflict:
                min_conflict = conflict
                ## 输出标准格式的result
                result = copy.deepcopy(self.schedules)
                ## 用于返回的conflict_list
                ret_conflict_list = []
                for conflict_detail in conflict_list:
                    res_item = {}
                    res_item[int(conflict_detail['class'][0:-1]) - 1] = value
                    ret_conflict_list.append(res_item)
                ## 将排课结果输出
                for cla in range(len(result)):
                    cla_name = self.classes_name[cla]
                    for week in range(len(result[cla])):
                        for section in range(len(result[cla][week])):
                            course = result[cla][week][section]
                            if course is None or course not in self.class_course_teacher_info[cla + 1]:
                                continue
                            result[cla][week][section] = cla_name + '-' + result[cla][week][section] + '-' + \
                                                         self.class_course_teacher_info[cla + 1][course]
                self.output_schedules()
            if min_conflict == 0 or solve_conflict_flag:
                break
            iteration += 1
            self.morning_count_failed = 0
            count_failed = 0
            # 5/26 如果iteration达到Max还没有，输出result
            # ...
        print("min conflict number:{}".format(min_conflict))
        # 如果没有解决完所有的冲突：没有经过深度搜索的，返回最小冲突的结果 result 和 冲突课程位置 conflict_list 8/1
        # --如果深度搜索对课表进行了修改（不管是否完全解决），还需要一个函数，检测最后的结果中，有哪些冲突点，并返回冲突点集
        print(result, ret_conflict_list)
        return result, ret_conflict_list  # 没有冲突或是最小冲突的排课结果

    # datasource 格式转换
    @staticmethod
    def translate_from(datasource: data_source.DataSource):
        data = datasource
        # 对开课程
        # {'1班-化': {'class': '2班', 'subject': '生'}, '9班-政': {'class': '7班', 'subject': '地'}}
        data.LOCKED_CLA_COURSE = datasource.BIND_CLASSES
        classes_num = {}
        # classes_num ={“1班”：1，“2班”：2， ... ， “9班”：9}
        for i in range(len(data.CLASSES)):
            classes_num[data.CLASSES[i]] = i + 1
        data.CLASSES_NUM = classes_num
        # classes_course_teacher
        keys = list(data.CLASSES_COURSE_TEACHER.keys())
        # 把key从1班，2班，3班转化成 1，2，3
        for key in keys:
            new_key = classes_num[key]
            data.CLASSES_COURSE_TEACHER[new_key] = data.CLASSES_COURSE_TEACHER[key]
            data.CLASSES_COURSE_TEACHER.pop(key)
            # 5/26 新增数据结构 ========================================================
            data.COURSE_TOTAL[new_key] = data.COURSE_TOTAL[key]
            data.COURSE_TOTAL.pop(key)
            data.COURSE_LINKED_HOURS[new_key] = data.COURSE_LINKED_HOURS[key]
            data.COURSE_LINKED_HOURS.pop(key)
            data.COURSE_MORNING_TOTAL[new_key] = data.COURSE_MORNING_TOTAL[key]
            data.COURSE_MORNING_TOTAL.pop(key)
            # ========================================================================
        data.FLOW_SIGN = '走'
        data.NOT_ASSIGN_SIGN = '不排'
        data.CLASS_NUM = len(data.CLASSES)
        return data
